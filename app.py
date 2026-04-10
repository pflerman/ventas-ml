#!/usr/bin/env python3
"""Ventas PaliShopping - Lista ventas de MercadoLibre en vivo."""
import io
import json
import os
import re
import sys
import threading
import tkinter as tk
import unicodedata
import webbrowser
import calendar as _calendar
from datetime import datetime, date
from pathlib import Path
from tkinter import filedialog, font as tkfont
from tkinter import messagebox, ttk

from PIL import Image, ImageTk

import dolar
import frase
import local_store
import whatsapp_send

# WSL no sincroniza el clipboard de Tk (X11/WSLg) con el de Windows.
# Si estamos en WSL usamos clip.exe directo para que Ctrl+V funcione en apps Windows.
def _is_wsl() -> bool:
    try:
        return "microsoft" in Path("/proc/version").read_text().lower()
    except OSError:
        return False

_IS_WSL = _is_wsl()
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

USER_ID = 24192412
PAGE_SIZE = 50
API_URL = "https://api.mercadolibre.com/orders/search"

CHECKED = "☑"
UNCHECKED = "☐"

# Reutilizamos MLAuth del proyecto mercadolibre-mcp para manejar refresh.
ML_MCP_DIR = Path.home() / "Proyectos" / "mercadolibre-mcp"
CREDENTIALS_PATH = (
    Path.home() / "Proyectos" / "ml-scripts" / "config" / "ml_credentials_palishopping.json"
)
sys.path.insert(0, str(ML_MCP_DIR))
from ml_auth import MLAuth  # noqa: E402


def get_auth() -> MLAuth:
    return MLAuth(str(CREDENTIALS_PATH))


def fetch_orders(auth: MLAuth, offset: int = 0, limit: int = PAGE_SIZE) -> dict:
    params = {
        "seller": USER_ID,
        "sort": "date_desc",
        "offset": offset,
        "limit": limit,
    }
    url = f"{API_URL}?{urlencode(params)}"
    req = Request(url, headers={"Authorization": f"Bearer {auth.access_token}"})
    with urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8"))


def parse_iso(iso: str):
    if not iso:
        return None
    try:
        return datetime.fromisoformat(iso.replace("Z", "+00:00"))
    except ValueError:
        return None


def format_day(dt) -> str:
    return dt.strftime("%d/%m/%Y") if dt else ""


def format_time(dt) -> str:
    return dt.strftime("%H:%M") if dt else ""


def extract_sku(order_item: dict) -> str:
    """SKU puede venir en distintos campos según el item."""
    item = order_item.get("item") or {}
    # 1. seller_sku directo del item
    sku = item.get("seller_sku")
    if sku:
        return str(sku)
    # 2. seller_custom_field (legacy)
    sku = item.get("seller_custom_field")
    if sku:
        return str(sku)
    # 3. variation_attributes con SELLER_SKU
    for attr in item.get("variation_attributes") or []:
        if (attr.get("id") or "").upper() == "SELLER_SKU":
            v = attr.get("value_name")
            if v:
                return str(v)
    return ""


def format_price(value) -> str:
    try:
        n = float(value)
    except (TypeError, ValueError):
        return ""
    return f"${n:,.2f}"


def _normalize(text: str) -> str:
    """Minúsculas + sin acentos/tildes."""
    text = (text or "").lower()
    nfkd = unicodedata.normalize("NFD", text)
    return "".join(c for c in nfkd if unicodedata.category(c) != "Mn")


# Costeo de importación desde China.
# precio_fob (USD) -> nacionalizado (USD) -> en pesos -> + ganancia hermano
NACIONALIZACION_MULT = 1.9
GANANCIA_HERMANO_MULT = 1.30

# El neto MP NO se calcula desde la API — se carga a mano por venta desde el
# panel de detalle. Pablo va a Mercado Pago, copia el neto real, y lo pega en
# el modal. Persistido en local_store.neto_manual por order_id. Cualquier
# cálculo que dependa del neto (ganancia, margen, totales) lee de ahí.


class VentasApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Ventas PaliShopping")
        self.root.geometry("900x780")

        try:
            self.auth = get_auth()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar las credenciales:\n{e}")
            raise SystemExit(1)

        self.offset = 0
        self.total = 0
        self.loading = False
        self.row_to_order: dict = {}  # tree item id -> order_id (solo hojas)
        self.row_base: dict = {}  # tree item id -> "odd"/"even"
        self.leaf_to_item: dict = {}  # tree leaf id -> {item_id, variation_id, sku, title}
        self.day_nodes: dict = {}  # "DD/MM/YYYY" -> parent row id
        self.day_count: dict = {}  # "DD/MM/YYYY" -> int
        self.day_total: dict = {}  # "DD/MM/YYYY" -> float
        self._last_costo_unitario: float | None = None  # seteado por _render_costo
        # Para el filtro: por cada leaf cargado guardamos el día y el texto
        # normalizado para matchear sin tener que re-leer values del Treeview.
        self._leaves_meta: dict = {}  # leaf_id -> {"day_key", "row_text", "order"}
        self._leaf_order_counter = 0  # contador para preservar orden de inserción
        self._filter_active = False

        # Carga local del JSON (sync, instantánea).
        local_store.init()

        self._build_ui()
        self._cargar_dolar_async()
        self.refresh()

    def _build_ui(self):
        base_font = tkfont.nametofont("TkDefaultFont")
        base_font.configure(size=11)
        self.root.option_add("*Font", base_font)

        style = ttk.Style()
        style.configure("Treeview", rowheight=28, font=("TkDefaultFont", 11))
        style.configure("Treeview.Heading", font=("TkDefaultFont", 11, "bold"))

        # Tk 8.6 bug: tag backgrounds ignored a menos que se reescriba el style.map.
        def _fixed_map(option):
            return [
                e for e in style.map("Treeview", query_opt=option)
                if e[:2] != ("!disabled", "!selected")
            ]
        style.map(
            "Treeview",
            foreground=_fixed_map("foreground"),
            background=_fixed_map("background"),
        )

        # Notebook con dos pestañas: "Ventas" (todo lo de siempre) y
        # "Consolidados" (sección manual e independiente). El status_var y
        # dolar_var quedan en el bottom de la pestaña Ventas — _flash_status
        # sigue funcionando, sólo que cuando estás en Consolidados los
        # mensajes no se ven (limitación menor, asumida).
        self.notebook = ttk.Notebook(self.root, padding=6)
        self.notebook.pack(fill="both", expand=True)

        container = ttk.Frame(self.notebook, padding=4)
        self.notebook.add(container, text="Ventas")

        consolidados_tab = ttk.Frame(self.notebook, padding=4)
        self.notebook.add(consolidados_tab, text="Consolidados")
        self._consolidados_tab = consolidados_tab

        liquidacion_tab = ttk.Frame(self.notebook, padding=4)
        self.notebook.add(liquidacion_tab, text="Liquidación")
        self._liquidacion_tab = liquidacion_tab

        paned = ttk.PanedWindow(container, orient="horizontal")
        paned.pack(fill="both", expand=True)

        tree_frame = ttk.Frame(paned)
        paned.add(tree_frame, weight=3)

        # ─── Filter bar ───
        filter_bar = ttk.Frame(tree_frame)
        filter_bar.pack(side="top", fill="x", pady=(0, 6))

        ttk.Label(filter_bar, text="🔍 Buscar:").pack(side="left", padx=(0, 4))
        self._buscar_var = tk.StringVar()
        self._buscar_var.trace_add("write", lambda *_: self._refresh_tree_filter())
        self._buscar_entry = ttk.Entry(
            filter_bar, textvariable=self._buscar_var, width=24
        )
        self._buscar_entry.pack(side="left", padx=(0, 12))

        ttk.Label(filter_bar, text="✕ Excluir:").pack(side="left", padx=(0, 4))
        self._excluir_var = tk.StringVar()
        self._excluir_var.trace_add("write", lambda *_: self._refresh_tree_filter())
        self._excluir_entry = ttk.Entry(
            filter_bar, textvariable=self._excluir_var, width=24
        )
        self._excluir_entry.pack(side="left", padx=(0, 8))

        self._BUSCAR_PLACEHOLDER = "SKU o producto..."
        self._EXCLUIR_PLACEHOLDER = "palabras a excluir..."
        self._setup_filter_placeholder(
            self._buscar_entry, self._buscar_var, self._BUSCAR_PLACEHOLDER
        )
        self._setup_filter_placeholder(
            self._excluir_entry, self._excluir_var, self._EXCLUIR_PLACEHOLDER
        )

        ttk.Button(
            filter_bar, text="Limpiar", command=self._limpiar_filtros
        ).pack(side="left")

        # Toggle "solo con nota" — filtra el tree para mostrar únicamente
        # ventas que tienen una nota asociada (casos que el usuario marcó
        # como "sospechosos para revisar después").
        self._solo_con_nota_var = tk.BooleanVar(value=False)
        self._solo_con_nota_chk = ttk.Checkbutton(
            filter_bar,
            text="📝 Solo con nota",
            variable=self._solo_con_nota_var,
            command=self._refresh_tree_filter,
        )
        self._solo_con_nota_chk.pack(side="left", padx=(12, 0))

        # Panel lateral scrollable: PanedWindow ⊃ outer ⊃ Canvas + Scrollbar,
        # con `detail_frame` viviendo adentro del Canvas via create_window.
        # El resto del código sigue manipulando self.detail_frame sin enterarse.
        detail_outer = ttk.Frame(paned)
        paned.add(detail_outer, weight=1)

        self.detail_canvas = tk.Canvas(
            detail_outer, highlightthickness=0, borderwidth=0
        )
        detail_vsb = ttk.Scrollbar(
            detail_outer, orient="vertical", command=self.detail_canvas.yview
        )
        self.detail_canvas.configure(yscrollcommand=detail_vsb.set)
        detail_vsb.pack(side="right", fill="y")
        self.detail_canvas.pack(side="left", fill="both", expand=True)

        self.detail_frame = ttk.Frame(self.detail_canvas, padding=(10, 4))
        self._detail_window_id = self.detail_canvas.create_window(
            (0, 0), window=self.detail_frame, anchor="nw"
        )

        # Cuando cambia el contenido del panel, actualizar el área scrollable.
        self.detail_frame.bind(
            "<Configure>",
            lambda e: self.detail_canvas.configure(
                scrollregion=self.detail_canvas.bbox("all")
            ),
        )
        # Cuando se redimensiona el Canvas (drag del PanedWindow / resize ventana),
        # forzar al frame interno a tener el mismo ancho — si no, los wraplength
        # se calculan sobre un ancho equivocado.
        self.detail_canvas.bind(
            "<Configure>",
            lambda e: self.detail_canvas.itemconfig(
                self._detail_window_id, width=e.width
            ),
        )

        # Mouse wheel sobre el panel: en Linux son Button-4/Button-5,
        # en Windows/macOS es <MouseWheel>. Solo activo el binding cuando
        # el cursor está sobre el panel para no robarle el wheel al Treeview.
        def _on_mousewheel(event):
            if event.num == 4 or getattr(event, "delta", 0) > 0:
                self.detail_canvas.yview_scroll(-3, "units")
            else:
                self.detail_canvas.yview_scroll(3, "units")

        def _bind_wheel(_e):
            self.detail_canvas.bind_all("<MouseWheel>", _on_mousewheel)
            self.detail_canvas.bind_all("<Button-4>", _on_mousewheel)
            self.detail_canvas.bind_all("<Button-5>", _on_mousewheel)

        def _unbind_wheel(_e):
            self.detail_canvas.unbind_all("<MouseWheel>")
            self.detail_canvas.unbind_all("<Button-4>")
            self.detail_canvas.unbind_all("<Button-5>")

        self.detail_canvas.bind("<Enter>", _bind_wheel)
        self.detail_canvas.bind("<Leave>", _unbind_wheel)

        self._build_detail_panel()

        columns = ("check", "fecha", "sku", "cant", "producto", "precio", "subtotal")
        self.tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="tree headings",
            selectmode="browse",
        )
        self.tree.heading("#0", text="Día")
        self.tree.heading("check", text=UNCHECKED, command=self._toggle_all)
        self.tree.heading("fecha", text="Hora")
        self.tree.heading("sku", text="SKU")
        self.tree.heading("cant", text="Cant")
        self.tree.heading("producto", text="Producto")
        self.tree.heading("precio", text="Precio U.")
        self.tree.heading("subtotal", text="Subtotal")
        self.tree.column("#0", width=200, anchor="w", stretch=False)
        self.tree.column("check", width=40, anchor="center", stretch=False)
        self.tree.column("fecha", width=60, anchor="w", stretch=False)
        self.tree.column("sku", width=90, anchor="w", stretch=False)
        self.tree.column("cant", width=50, anchor="center", stretch=False)
        self.tree.column("producto", width=280, anchor="w")
        self.tree.column("precio", width=110, anchor="e", stretch=False)
        self.tree.column("subtotal", width=110, anchor="e", stretch=False)
        self.tree.bind("<Button-1>", self._on_click)
        self.tree.bind("<Button-3>", self._on_right_click)
        self.tree.bind("<Double-Button-1>", self._on_double_click)
        self.tree.bind("<Control-Button-1>", self._on_ctrl_click_fob)
        self.tree.bind("<Alt-Button-1>", self._on_alt_click_mult)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        self._modifiers_active: set[str] = set()  # subset of {"ctrl", "shift", "alt"}
        self._status_before_hint = ""
        self.root.bind("<KeyPress-Control_L>", lambda e: self._modifier_pressed("ctrl"))
        self.root.bind("<KeyPress-Control_R>", lambda e: self._modifier_pressed("ctrl"))
        self.root.bind("<KeyRelease-Control_L>", lambda e: self._modifier_released("ctrl"))
        self.root.bind("<KeyRelease-Control_R>", lambda e: self._modifier_released("ctrl"))
        self.root.bind("<KeyPress-Alt_L>", lambda e: self._modifier_pressed("alt"))
        self.root.bind("<KeyPress-Alt_R>", lambda e: self._modifier_pressed("alt"))
        self.root.bind("<KeyRelease-Alt_L>", lambda e: self._modifier_released("alt"))
        self.root.bind("<KeyRelease-Alt_R>", lambda e: self._modifier_released("alt"))

        # F-keys sobre la fila seleccionada (acciones rápidas de "ir al sitio").
        # F1: detalle de la venta en ML
        # F2: detalle del pago en Mercado Pago
        # F3: editor de la publicación en ML
        # F4: publicación pública (la que ve el comprador)
        # F6: abrir las cuatro de una
        self.root.bind("<F1>", self._on_f1_detalle_venta)
        self.root.bind("<F2>", self._on_f2_pago_mp)
        self.root.bind("<F3>", self._on_f3_edit_publicacion)
        self.root.bind("<F4>", self._on_f4_publi_publica)
        self.root.bind("<F6>", self._on_f6_abrir_todo)

        self.context_menu = tk.Menu(self.tree, tearoff=0)
        self.context_menu.add_command(
            label="Copiar SKU", command=self._copy_clicked_sku
        )
        self.context_menu.add_command(
            label="Copiar título", command=self._copy_clicked_title
        )
        self.context_menu.add_command(
            label="Copiar ID publicación", command=self._copy_clicked_item_id
        )
        self.context_menu.add_command(
            label="Copiar selección (WhatsApp)",
            command=self._copy_selected_to_clipboard,
        )
        self.context_menu.add_separator()
        self.context_menu.add_command(label="Refrescar todo", command=self.refresh)
        self.context_menu.bind("<FocusOut>", lambda e: self.context_menu.unpost())
        self._right_clicked_row = None

        self.tree.tag_configure("odd", background="#f5f5f5")
        self.tree.tag_configure("even", background="#ffffff")
        self.tree.tag_configure("selected", background="#C6EFCE")
        self.tree.tag_configure(
            "day", background="#dce6f0", font=("TkDefaultFont", 11, "bold")
        )
        # Filas con nota: foreground violeta (no toca el background así no
        # pisa al tag "selected" del check). Es una señal visual de que ese
        # caso lo marcaste para revisar después.
        self.tree.tag_configure("with_note", foreground="#7d3c98")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        bottom = ttk.Frame(container)
        bottom.pack(fill="x", pady=(10, 0))

        self.status_var = tk.StringVar(value="")
        ttk.Label(bottom, textvariable=self.status_var).pack(side="left")

        self.dolar_var = tk.StringVar(value="USD …")
        ttk.Label(
            bottom,
            textvariable=self.dolar_var,
            foreground="#1e7a1e",
            font=("TkDefaultFont", 10, "bold"),
        ).pack(side="left", padx=(16, 0))

        # Mini totalizador de ventas seleccionadas. Aparece solo cuando hay 1+.
        self.totales_costo_var = tk.StringVar(value="")
        self.totales_costo_lbl = tk.Label(
            bottom,
            textvariable=self.totales_costo_var,
            foreground="#c0392b",
            font=("TkDefaultFont", 10),
        )
        self.totales_costo_lbl.pack(side="left", padx=(16, 0))

        self.totales_ganancia_var = tk.StringVar(value="")
        self.totales_ganancia_lbl = tk.Label(
            bottom,
            textvariable=self.totales_ganancia_var,
            foreground="#1e7a1e",
            font=("TkDefaultFont", 10, "bold"),
        )
        self.totales_ganancia_lbl.pack(side="left", padx=(8, 0))

        self.btn_more = ttk.Button(bottom, text="Cargar más", command=self.load_more)
        self.btn_more.pack(side="right", padx=(6, 0))
        self.btn_refresh = ttk.Button(bottom, text="Actualizar", command=self.refresh)
        self.btn_refresh.pack(side="right")
        self.btn_export = ttk.Button(
            bottom, text="Exportar Excel", command=self.export_excel
        )
        self.btn_export.pack(side="right", padx=(0, 6))
        self.btn_totales = ttk.Button(
            bottom, text="Detalle totales", command=self._open_totales_modal
        )
        self.btn_totales.pack(side="right", padx=(0, 6))

        # Pestaña Consolidados (independiente de las ventas).
        self._build_consolidados_tab()

        # Pestaña Liquidación (calendario + links MP por día).
        self._build_liquidacion_tab()

    def _build_detail_panel(self):
        ttk.Label(
            self.detail_frame,
            text="Producto",
            font=("TkDefaultFont", 12, "bold"),
        ).pack(anchor="w", pady=(0, 8))

        self.detail_status_var = tk.StringVar(value="Seleccioná una venta")
        self.detail_status_lbl = ttk.Label(
            self.detail_frame,
            textvariable=self.detail_status_var,
            foreground="#888",
            wraplength=240,
            justify="left",
        )
        self.detail_status_lbl.pack(anchor="w", pady=(0, 8))

        ttk.Label(
            self.detail_frame,
            text="Etiquetas",
            font=("TkDefaultFont", 10, "bold"),
        ).pack(anchor="w", pady=(8, 4))

        # Chips de etiquetas asignadas al SKU actual.
        self.detail_tags_frame = ttk.Frame(self.detail_frame)
        self.detail_tags_frame.pack(fill="x", anchor="w")

        # Combobox para elegir/agregar una etiqueta del catálogo + botón "+"
        # para crear una etiqueta nueva al catálogo. SKU actual cacheado para
        # que los handlers sepan a qué venta aplicar.
        self._detail_current_sku: str | None = None
        tag_picker = ttk.Frame(self.detail_frame)
        tag_picker.pack(fill="x", anchor="w", pady=(4, 0))
        self._tag_combo_var = tk.StringVar()
        self._tag_combo = ttk.Combobox(
            tag_picker,
            textvariable=self._tag_combo_var,
            state="readonly",
            width=18,
        )
        self._tag_combo.pack(side="left", padx=(0, 4))
        ttk.Button(
            tag_picker, text="Agregar", width=8, command=self._on_add_tag_click
        ).pack(side="left")
        ttk.Button(
            tag_picker, text="+", width=2, command=self._on_new_tag_click
        ).pack(side="left", padx=(4, 0))

        ttk.Separator(self.detail_frame, orient="horizontal").pack(
            fill="x", pady=(12, 8)
        )

        ttk.Label(
            self.detail_frame,
            text="Costo importación",
            font=("TkDefaultFont", 10, "bold"),
        ).pack(anchor="w", pady=(0, 4))

        self.detail_costo_frame = ttk.Frame(self.detail_frame)
        self.detail_costo_frame.pack(fill="x", anchor="w")

        ttk.Separator(self.detail_frame, orient="horizontal").pack(
            fill="x", pady=(12, 8)
        )

        ttk.Label(
            self.detail_frame,
            text="Ganancia Andrés",
            font=("TkDefaultFont", 10, "bold"),
        ).pack(anchor="w", pady=(0, 4))

        self.detail_andres_frame = ttk.Frame(self.detail_frame)
        self.detail_andres_frame.pack(fill="x", anchor="w")

        ttk.Separator(self.detail_frame, orient="horizontal").pack(
            fill="x", pady=(12, 8)
        )

        ttk.Label(
            self.detail_frame,
            text="Cobro Mercado Pago",
            font=("TkDefaultFont", 10, "bold"),
        ).pack(anchor="w", pady=(0, 4))

        self.detail_payment_frame = ttk.Frame(self.detail_frame)
        self.detail_payment_frame.pack(fill="x", anchor="w")

        self.detail_payment_hint = ttk.Label(
            self.detail_frame,
            text="F2 → cobro en MP  ·  F1 venta  ·  F3 editor  ·  F4 publi  ·  F6 todas",
            foreground="#888",
            font=("TkDefaultFont", 9, "italic"),
        )
        self.detail_payment_hint.pack(anchor="w", pady=(8, 0))

        ttk.Separator(self.detail_frame, orient="horizontal").pack(
            fill="x", pady=(12, 8)
        )

        ttk.Label(
            self.detail_frame,
            text="Ganancia total",
            font=("TkDefaultFont", 10, "bold"),
        ).pack(anchor="w", pady=(0, 4))

        self.detail_ganancia_frame = ttk.Frame(self.detail_frame)
        self.detail_ganancia_frame.pack(fill="x", anchor="w")

        # ─── Nota libre por venta ───
        # Persistida en local_store (data.json) por order_id.
        # Autosave on focus-out, sync (escritura local instantánea).
        ttk.Separator(self.detail_frame, orient="horizontal").pack(
            fill="x", pady=(16, 6)
        )
        ttk.Label(
            self.detail_frame,
            text="📝 Nota",
            font=("TkDefaultFont", 10, "bold"),
            foreground="#7d3c98",
        ).pack(anchor="w")
        self.detail_nota_text = tk.Text(
            self.detail_frame,
            height=4,
            wrap="word",
            font=("TkDefaultFont", 10),
            relief="solid",
            borderwidth=1,
            highlightthickness=0,
        )
        self.detail_nota_text.pack(fill="x", anchor="w", pady=(2, 0))
        # State para no pisar la nota cuando recién cargás otra venta
        # (la primer carga es _load_nota_into_widget, no input del usuario).
        self._nota_loading = False
        self._nota_current_order: str | None = None
        self.detail_nota_text.bind("<FocusOut>", self._on_nota_focus_out)
        ttk.Label(
            self.detail_frame,
            text="Se guarda automáticamente al cambiar de venta.",
            foreground="#888",
            font=("TkDefaultFont", 9, "italic"),
        ).pack(anchor="w", pady=(2, 0))

        ttk.Separator(self.detail_frame, orient="horizontal").pack(
            fill="x", pady=(16, 8)
        )
        botones_frame = ttk.Frame(self.detail_frame)
        botones_frame.pack(fill="x", anchor="w")
        ttk.Button(
            botones_frame,
            text="✨ Frase del día ✨",
            command=self._open_frase_modal,
        ).pack(side="left")
        ttk.Button(
            botones_frame,
            text="📋 Copiar informe",
            command=self._copy_informe_wasap,
        ).pack(side="left", padx=(8, 0))
        ttk.Button(
            botones_frame,
            text="📄 Copiar informe lite",
            command=self._copy_informe_lite,
        ).pack(side="left", padx=(8, 0))

    # ──────────────────── Pestaña Consolidados ────────────────────
    # Sección 100% manual e independiente. Cada tarjeta es una "consolidación
    # de pago" entre Pablo y Andrés. Persiste en data.json bajo "consolidados"
    # y se sube al repo como el resto de los datos.

    CONSOLIDADOS_COLS = 3  # cards por fila en el grid

    def _build_consolidados_tab(self):
        tab = self._consolidados_tab

        # ─── Top bar: nuevo + filtro ───
        topbar = ttk.Frame(tab)
        topbar.pack(fill="x", pady=(0, 8))

        ttk.Button(
            topbar,
            text="➕ Nuevo consolidado",
            command=lambda: self._open_consolidado_modal(None),
        ).pack(side="left")

        self._consol_solo_activos = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            topbar,
            text="Solo activos",
            variable=self._consol_solo_activos,
            command=self._refresh_consolidados,
        ).pack(side="left", padx=(12, 0))

        ttk.Label(
            topbar,
            text="(click en una tarjeta para editar)",
            foreground="#888",
            font=("TkDefaultFont", 9, "italic"),
        ).pack(side="left", padx=(16, 0))

        # ─── Área scrolleable con los cards ───
        # Canvas + Frame interno + Scrollbar — patrón clásico de Tk para
        # listas de widgets de altura variable.
        canvas_outer = ttk.Frame(tab)
        canvas_outer.pack(fill="both", expand=True)

        self._consol_canvas = tk.Canvas(
            canvas_outer, highlightthickness=0, borderwidth=0
        )
        consol_vsb = ttk.Scrollbar(
            canvas_outer, orient="vertical", command=self._consol_canvas.yview
        )
        self._consol_canvas.configure(yscrollcommand=consol_vsb.set)
        consol_vsb.pack(side="right", fill="y")
        self._consol_canvas.pack(side="left", fill="both", expand=True)

        self._consol_grid_frame = ttk.Frame(self._consol_canvas)
        self._consol_grid_window = self._consol_canvas.create_window(
            (0, 0), window=self._consol_grid_frame, anchor="nw"
        )

        self._consol_grid_frame.bind(
            "<Configure>",
            lambda e: self._consol_canvas.configure(
                scrollregion=self._consol_canvas.bbox("all")
            ),
        )
        self._consol_canvas.bind(
            "<Configure>",
            lambda e: self._consol_canvas.itemconfig(
                self._consol_grid_window, width=e.width
            ),
        )

        # Mouse wheel sobre el canvas (Linux usa Button-4/5).
        def _on_wheel(event):
            if event.num == 4 or getattr(event, "delta", 0) > 0:
                self._consol_canvas.yview_scroll(-3, "units")
            else:
                self._consol_canvas.yview_scroll(3, "units")

        def _bind_wheel(_e):
            self._consol_canvas.bind_all("<MouseWheel>", _on_wheel)
            self._consol_canvas.bind_all("<Button-4>", _on_wheel)
            self._consol_canvas.bind_all("<Button-5>", _on_wheel)

        def _unbind_wheel(_e):
            self._consol_canvas.unbind_all("<MouseWheel>")
            self._consol_canvas.unbind_all("<Button-4>")
            self._consol_canvas.unbind_all("<Button-5>")

        self._consol_canvas.bind("<Enter>", _bind_wheel)
        self._consol_canvas.bind("<Leave>", _unbind_wheel)

        self._refresh_consolidados()

    def _refresh_consolidados(self):
        """Reconstruye el grid de cards desde local_store. Newest top-right,
        fillea derecha→izquierda, wrap a la fila siguiente."""
        # Limpiar contenido previo
        for w in self._consol_grid_frame.winfo_children():
            w.destroy()

        consolidados = local_store.list_consolidados()
        if self._consol_solo_activos.get():
            consolidados = [c for c in consolidados if c.get("activo", True)]

        # Newest first: orden por fecha_creacion desc, fallback por orden inverso
        # de inserción (los más nuevos del array primero).
        consolidados.sort(
            key=lambda c: (c.get("fecha_creacion") or ""), reverse=True
        )

        if not consolidados:
            ttk.Label(
                self._consol_grid_frame,
                text="(no hay consolidados — clickeá '➕ Nuevo consolidado' para empezar)",
                foreground="#888",
                font=("TkDefaultFont", 10, "italic"),
                padding=20,
            ).pack(anchor="center")
            return

        cols = self.CONSOLIDADOS_COLS
        for i, c in enumerate(consolidados):
            row = i // cols
            # Right-to-left fill: el primero (más nuevo) va a la derecha.
            col = (cols - 1) - (i % cols)
            card = self._make_consolidado_card(self._consol_grid_frame, c)
            card.grid(
                row=row, column=col, padx=8, pady=8, sticky="nsew"
            )
        # Que las columnas se distribuyan equitativamente.
        for col in range(cols):
            self._consol_grid_frame.grid_columnconfigure(col, weight=1, uniform="cols")

    def _make_consolidado_card(self, parent, c: dict) -> tk.Widget:
        """Construye un card visual para un consolidado dado. Retorna el frame."""
        # Estilo: borde sólido, fondo blanco si activo, gris si inactivo.
        bg = "#ffffff" if c.get("activo", True) else "#ececec"
        card = tk.Frame(
            parent,
            relief="solid",
            borderwidth=1,
            background=bg,
            padx=10,
            pady=10,
        )

        cid = c.get("id", "")
        monto_deuda = float(c.get("monto_deuda") or 0)
        credito = float(c.get("credito") or 0)
        neto = monto_deuda - credito
        activo = c.get("activo", True)
        facturado = c.get("facturado", False)

        # ─── Header: fecha creación + botón ✕ ───
        header = tk.Frame(card, background=bg)
        header.pack(fill="x")
        tk.Label(
            header,
            text=f"📅 Creado: {c.get('fecha_creacion') or '—'}",
            background=bg,
            font=("TkDefaultFont", 10, "bold"),
        ).pack(side="left")
        tk.Button(
            header,
            text="✕",
            command=lambda i=cid: self._delete_consolidado(i),
            relief="flat",
            background=bg,
            foreground="#c0392b",
            borderwidth=0,
            cursor="hand2",
            font=("TkDefaultFont", 11, "bold"),
        ).pack(side="right")

        tk.Frame(card, background="#bbb", height=1).pack(fill="x", pady=(6, 6))

        # ─── Período ───
        periodo_txt = (
            f"📆 Período: {c.get('fecha_desde') or '—'}  →  {c.get('fecha_hasta') or '—'}"
        )
        tk.Label(
            card,
            text=periodo_txt,
            background=bg,
            anchor="w",
            justify="left",
            font=("TkDefaultFont", 10),
        ).pack(fill="x", pady=(0, 4))

        # ─── Fecha de pago (RESALTADA en rojo si activo) ───
        pago_txt = c.get("fecha_pago") or "(sin fecha)"
        pago_lbl = tk.Label(
            card,
            text=f"💸 Fecha de pago: {pago_txt}",
            background=bg,
            foreground="#c0392b" if activo else "#666",
            anchor="w",
            justify="left",
            font=("TkDefaultFont", 11, "bold"),
        )
        pago_lbl.pack(fill="x", pady=(0, 6))

        tk.Frame(card, background="#bbb", height=1).pack(fill="x", pady=(0, 6))

        # ─── Montos ───
        def money_row(label, value, color):
            r = tk.Frame(card, background=bg)
            r.pack(fill="x", pady=1)
            tk.Label(
                r, text=label, background=bg, font=("TkDefaultFont", 10)
            ).pack(side="left")
            tk.Label(
                r,
                text=format_price(value),
                background=bg,
                foreground=color,
                font=("TkDefaultFont", 10, "bold"),
            ).pack(side="right")

        money_row("Le debo a Andrés:", monto_deuda, "#c0392b")
        if credito:
            money_row("− Crédito (Andrés→yo):", credito, "#1e7a1e")

        tk.Frame(card, background="#bbb", height=1).pack(fill="x", pady=(6, 4))

        neto_row = tk.Frame(card, background=bg)
        neto_row.pack(fill="x", pady=(2, 6))
        tk.Label(
            neto_row,
            text="Neto a pagar:",
            background=bg,
            font=("TkDefaultFont", 11, "bold"),
        ).pack(side="left")
        tk.Label(
            neto_row,
            text=format_price(neto),
            background=bg,
            foreground="#1f4e9d",
            font=("TkDefaultFont", 12, "bold", "underline"),
        ).pack(side="right")

        # ─── Nota libre ───
        nota_txt = c.get("nota") or ""
        if nota_txt:
            tk.Label(
                card,
                text=f"📝 {nota_txt}",
                background=bg,
                foreground="#7d3c98",
                wraplength=240,
                justify="left",
                anchor="w",
                font=("TkDefaultFont", 9, "italic"),
            ).pack(fill="x", pady=(2, 6))

        # ─── Checks: Activo + Facturado ───
        checks_row = tk.Frame(card, background=bg)
        checks_row.pack(fill="x", pady=(4, 0))

        activo_var = tk.BooleanVar(value=activo)
        def toggle_activo(i=cid, v=activo_var):
            local_store.update_consolidado(i, activo=v.get())
            self._refresh_consolidados()
        tk.Checkbutton(
            checks_row,
            text="✅ Activo",
            variable=activo_var,
            command=toggle_activo,
            background=bg,
            activebackground=bg,
            font=("TkDefaultFont", 9),
        ).pack(side="left")

        fact_var = tk.BooleanVar(value=facturado)
        def toggle_fact(i=cid, v=fact_var):
            local_store.update_consolidado(i, facturado=v.get())
            self._refresh_consolidados()
        tk.Checkbutton(
            checks_row,
            text="🧾 Facturado",
            variable=fact_var,
            command=toggle_fact,
            background=bg,
            activebackground=bg,
            font=("TkDefaultFont", 9),
        ).pack(side="left", padx=(8, 0))

        # Click en el card (en cualquier parte que no sea botón/check) → editar.
        def open_edit(_e=None, i=cid):
            self._open_consolidado_modal(i)

        # Bindear a card y sus children "estáticos" (no a botones/checks).
        for w in (card, header, pago_lbl, neto_row):
            w.bind("<Button-1>", open_edit)
        card.configure(cursor="hand2")

        return card

    def _delete_consolidado(self, cid: str):
        if not cid:
            return
        c = local_store.get_consolidado(cid)
        if not c:
            return
        # Confirmación porque borrar es destructivo
        ok = messagebox.askyesno(
            "Borrar consolidado",
            f"¿Borrar el consolidado del {c.get('fecha_creacion') or '—'}?\n"
            f"Le debo a Andrés: {format_price(c.get('monto_deuda') or 0)}\n\n"
            "Esto no se puede deshacer.",
            parent=self.root,
        )
        if not ok:
            return
        local_store.delete_consolidado(cid)
        self._refresh_consolidados()

    def _open_consolidado_modal(self, cid: str | None):
        """Modal de alta o edición. Si cid es None es alta nueva."""
        existing = local_store.get_consolidado(cid) if cid else None

        win = tk.Toplevel(self.root)
        win.title("Editar consolidado" if existing else "Nuevo consolidado")
        win.transient(self.root)
        win.resizable(False, False)

        frame = ttk.Frame(win, padding=15)
        frame.pack(fill="both", expand=True)

        ttk.Label(
            frame,
            text="Editar consolidado" if existing else "Nuevo consolidado",
            font=("TkDefaultFont", 11, "bold"),
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 12))

        # Campos: cada uno (label, entry).
        # Las fechas son texto libre (formato YYYY-MM-DD recomendado pero
        # sin validación rígida — Pablo escribe lo que quiera).
        labels = [
            ("Fecha creación", "fecha_creacion"),
            ("Fecha desde", "fecha_desde"),
            ("Fecha hasta", "fecha_hasta"),
            ("Fecha de pago", "fecha_pago"),
            ("Le debo a Andrés ($)", "monto_deuda"),
            ("Crédito Andrés→yo ($)", "credito"),
            ("Nota", "nota"),
        ]
        vars_map: dict[str, tk.StringVar] = {}
        for i, (lbl, key) in enumerate(labels, start=1):
            ttk.Label(frame, text=lbl + ":").grid(
                row=i, column=0, sticky="w", padx=(0, 8), pady=2
            )
            initial = ""
            if existing:
                v = existing.get(key)
                if v is not None and v != "":
                    if isinstance(v, float):
                        initial = f"{v:g}"
                    else:
                        initial = str(v)
            var = tk.StringVar(value=initial)
            vars_map[key] = var
            ttk.Entry(frame, textvariable=var, width=28).grid(
                row=i, column=1, sticky="ew", pady=2
            )

        # Checks
        activo_var = tk.BooleanVar(
            value=bool(existing.get("activo", True)) if existing else True
        )
        ttk.Checkbutton(
            frame, text="✅ Activo", variable=activo_var
        ).grid(row=len(labels) + 1, column=0, columnspan=2, sticky="w", pady=(8, 0))

        facturado_var = tk.BooleanVar(
            value=bool(existing.get("facturado", False)) if existing else False
        )
        ttk.Checkbutton(
            frame, text="🧾 Facturado", variable=facturado_var
        ).grid(row=len(labels) + 2, column=0, columnspan=2, sticky="w")

        btns = ttk.Frame(frame)
        btns.grid(row=len(labels) + 3, column=0, columnspan=2, sticky="ew", pady=(12, 0))

        def parse_money(raw: str) -> float:
            raw = (raw or "").strip().replace(",", ".")
            if raw == "":
                return 0.0
            return float(raw)

        def do_save():
            try:
                data = {
                    "fecha_creacion": vars_map["fecha_creacion"].get().strip(),
                    "fecha_desde": vars_map["fecha_desde"].get().strip(),
                    "fecha_hasta": vars_map["fecha_hasta"].get().strip(),
                    "fecha_pago": vars_map["fecha_pago"].get().strip(),
                    "monto_deuda": parse_money(vars_map["monto_deuda"].get()),
                    "credito": parse_money(vars_map["credito"].get()),
                    "nota": vars_map["nota"].get().strip(),
                    "activo": activo_var.get(),
                    "facturado": facturado_var.get(),
                }
            except ValueError:
                messagebox.showerror(
                    "Error", "Los montos deben ser números válidos.", parent=win
                )
                return

            try:
                if existing:
                    local_store.update_consolidado(existing["id"], **data)
                else:
                    local_store.add_consolidado(data)
            except Exception as e:
                messagebox.showerror(
                    "Error", f"No se pudo guardar:\n{e}", parent=win
                )
                return

            win.destroy()
            self._refresh_consolidados()
            self._flash_status("Consolidado guardado ✓")

        ttk.Button(btns, text="Guardar", command=do_save).pack(side="right")
        ttk.Button(btns, text="Cancelar", command=win.destroy).pack(
            side="right", padx=(0, 6)
        )

        win.bind("<Return>", lambda _e: do_save())
        win.bind("<KP_Enter>", lambda _e: do_save())
        win.bind("<Escape>", lambda _e: win.destroy())

        win.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() - win.winfo_width()) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() - win.winfo_height()) // 3
        win.geometry(f"+{x}+{y}")
        win.grab_set()

    # ──────────────────── Pestaña Liquidación ────────────────────
    # Calendario mensual + panel con links de Mercado Pago por día.
    # Todo manual: el usuario navega al mes, clickea un día, pega un link
    # y se persiste. Cuando vuelve a abrir la app sobre ese día, los links
    # aparecen listos para abrir/copiar.

    LIQ_DIAS_SEMANA = ["Lu", "Ma", "Mi", "Ju", "Vi", "Sá", "Do"]
    LIQ_MESES = [
        "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]
    LIQ_DIAS_NOMBRE = [
        "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo",
    ]

    def _build_liquidacion_tab(self):
        tab = self._liquidacion_tab

        today = date.today()
        self._liq_year = today.year
        self._liq_month = today.month
        self._liq_selected_date = today.isoformat()

        # ─── Banner "HOY" si tiene links ───
        # Reminder visual permanente: incluso si navegás a otro mes, este
        # banner te dice si hoy tenés links pendientes. Lo packeo PRIMERO
        # para que quede arriba de todo, y _liq_refresh_today_banner lo
        # esconde con pack_forget si hoy no hay links.
        self._liq_today_banner = tk.Frame(tab, background="#fff8c2")
        self._liq_today_banner_lbl = tk.Label(
            self._liq_today_banner,
            text="",
            background="#fff8c2",
            foreground="#8a6d00",
            font=("TkDefaultFont", 11, "bold"),
            padx=10,
            pady=6,
        )
        self._liq_today_banner_lbl.pack(side="left")
        tk.Button(
            self._liq_today_banner,
            text="Ver hoy →",
            command=self._liq_go_today,
            background="#fff8c2",
            relief="flat",
            cursor="hand2",
            font=("TkDefaultFont", 10, "bold"),
            foreground="#1f4e9d",
            borderwidth=0,
        ).pack(side="right", padx=(0, 10))
        # Pack inicial: visible. _liq_refresh_today_banner lo esconde si hace falta.
        self._liq_today_banner.pack(fill="x", pady=(0, 8))

        # ─── Top bar: nav del mes + botón hoy ───
        topbar = ttk.Frame(tab)
        topbar.pack(fill="x", pady=(0, 8))

        ttk.Button(
            topbar, text="◀", command=self._liq_prev_month, width=3
        ).pack(side="left")

        self._liq_month_label_var = tk.StringVar()
        tk.Label(
            topbar,
            textvariable=self._liq_month_label_var,
            font=("TkDefaultFont", 13, "bold"),
            foreground="#1a3a5c",
        ).pack(side="left", padx=(8, 8))

        ttk.Button(
            topbar, text="▶", command=self._liq_next_month, width=3
        ).pack(side="left")

        ttk.Button(
            topbar, text="📅 Hoy", command=self._liq_go_today
        ).pack(side="left", padx=(16, 0))

        # ─── Body split: calendario | panel de links ───
        body = ttk.PanedWindow(tab, orient="horizontal")
        body.pack(fill="both", expand=True)

        # Lado izquierdo: header de días + grid de días
        left = ttk.Frame(body)
        body.add(left, weight=2)

        dow_frame = ttk.Frame(left)
        dow_frame.pack(fill="x")
        for i, dow in enumerate(self.LIQ_DIAS_SEMANA):
            tk.Label(
                dow_frame,
                text=dow,
                font=("TkDefaultFont", 10, "bold"),
                foreground="#666",
                anchor="center",
            ).grid(row=0, column=i, sticky="nsew", padx=1, pady=(0, 4))
        for i in range(7):
            dow_frame.grid_columnconfigure(i, weight=1, uniform="liq_dow")

        self._liq_grid = ttk.Frame(left)
        self._liq_grid.pack(fill="both", expand=True)

        # Lado derecho: panel con links del día seleccionado
        right_outer = ttk.Frame(body)
        body.add(right_outer, weight=3)

        # Panel scrolleable
        self._liq_right_canvas = tk.Canvas(
            right_outer, highlightthickness=0, borderwidth=0
        )
        liq_vsb = ttk.Scrollbar(
            right_outer, orient="vertical", command=self._liq_right_canvas.yview
        )
        self._liq_right_canvas.configure(yscrollcommand=liq_vsb.set)
        liq_vsb.pack(side="right", fill="y")
        self._liq_right_canvas.pack(side="left", fill="both", expand=True)

        self._liq_right_frame = ttk.Frame(self._liq_right_canvas, padding=12)
        self._liq_right_window = self._liq_right_canvas.create_window(
            (0, 0), window=self._liq_right_frame, anchor="nw"
        )
        self._liq_right_frame.bind(
            "<Configure>",
            lambda e: self._liq_right_canvas.configure(
                scrollregion=self._liq_right_canvas.bbox("all")
            ),
        )
        self._liq_right_canvas.bind(
            "<Configure>",
            lambda e: self._liq_right_canvas.itemconfig(
                self._liq_right_window, width=e.width
            ),
        )

        # Wheel scroll sobre el panel derecho
        def _on_wheel(event):
            if event.num == 4 or getattr(event, "delta", 0) > 0:
                self._liq_right_canvas.yview_scroll(-3, "units")
            else:
                self._liq_right_canvas.yview_scroll(3, "units")

        def _bind_wheel(_e):
            self._liq_right_canvas.bind_all("<MouseWheel>", _on_wheel)
            self._liq_right_canvas.bind_all("<Button-4>", _on_wheel)
            self._liq_right_canvas.bind_all("<Button-5>", _on_wheel)

        def _unbind_wheel(_e):
            self._liq_right_canvas.unbind_all("<MouseWheel>")
            self._liq_right_canvas.unbind_all("<Button-4>")
            self._liq_right_canvas.unbind_all("<Button-5>")

        self._liq_right_canvas.bind("<Enter>", _bind_wheel)
        self._liq_right_canvas.bind("<Leave>", _unbind_wheel)

        self._liq_refresh_all()

    def _liq_refresh_all(self):
        self._liq_render_month_label()
        self._liq_refresh_today_banner()
        self._liq_render_calendar()
        self._liq_render_right_panel()

    def _liq_render_month_label(self):
        self._liq_month_label_var.set(
            f"{self.LIQ_MESES[self._liq_month]} {self._liq_year}"
        )

    def _liq_refresh_today_banner(self):
        today_iso = date.today().isoformat()
        n = local_store.count_links_dia(today_iso)
        if n > 0:
            self._liq_today_banner_lbl.config(
                text=f"🔔 Hoy ({today_iso}) tenés {n} link{'s' if n != 1 else ''} de liquidación"
            )
            if not self._liq_today_banner.winfo_ismapped():
                # Reusar pack original (arriba de todo, side default = top).
                self._liq_today_banner.pack(
                    fill="x", pady=(0, 8), side="top", before=self._liq_today_banner.master.winfo_children()[1]
                )
        else:
            if self._liq_today_banner.winfo_ismapped():
                self._liq_today_banner.pack_forget()

    def _liq_render_calendar(self):
        for w in self._liq_grid.winfo_children():
            w.destroy()

        cal = _calendar.Calendar(firstweekday=0)  # 0 = lunes
        weeks = cal.monthdayscalendar(self._liq_year, self._liq_month)
        today_iso = date.today().isoformat()
        dias_con = local_store.dias_con_links()

        for row, week in enumerate(weeks):
            for col, day in enumerate(week):
                if day == 0:
                    # Celda vacía (días del mes anterior/siguiente).
                    tk.Label(
                        self._liq_grid, text="", background="#f8f8f8"
                    ).grid(row=row, column=col, padx=1, pady=1, sticky="nsew")
                    continue

                fecha = f"{self._liq_year}-{self._liq_month:02d}-{day:02d}"
                is_today = (fecha == today_iso)
                is_selected = (fecha == self._liq_selected_date)
                has_links = fecha in dias_con
                count = local_store.count_links_dia(fecha) if has_links else 0

                # Estilo en cascada: seleccionado pisa hoy pisa con-links pisa default.
                if is_selected:
                    bg = "#7d3c98"
                    fg = "white"
                    border = 2
                elif is_today:
                    bg = "#fff8c2"
                    fg = "#8a6d00"
                    border = 2
                elif has_links:
                    bg = "#d4f4ef"  # teal claro
                    fg = "#117a65"
                    border = 1
                else:
                    bg = "#ffffff"
                    fg = "#000"
                    border = 1

                txt = str(day)
                if count > 0:
                    txt = f"{day}\n● {count}"

                btn = tk.Button(
                    self._liq_grid,
                    text=txt,
                    background=bg,
                    foreground=fg,
                    activebackground=bg,
                    relief="solid",
                    borderwidth=border,
                    font=("TkDefaultFont", 10, "bold"),
                    command=lambda d=fecha: self._liq_select_day(d),
                    cursor="hand2",
                    height=2,
                )
                btn.grid(row=row, column=col, padx=1, pady=1, sticky="nsew")

        for col in range(7):
            self._liq_grid.grid_columnconfigure(col, weight=1, uniform="liq_col")
        for row in range(len(weeks)):
            self._liq_grid.grid_rowconfigure(row, weight=1, uniform="liq_row")

    def _liq_render_right_panel(self):
        for w in self._liq_right_frame.winfo_children():
            w.destroy()

        fecha = self._liq_selected_date
        try:
            dt = datetime.strptime(fecha, "%Y-%m-%d")
            header_text = (
                f"{self.LIQ_DIAS_NOMBRE[dt.weekday()]}, "
                f"{dt.day} de {self.LIQ_MESES[dt.month]} de {dt.year}"
            )
        except Exception:
            header_text = fecha

        tk.Label(
            self._liq_right_frame,
            text=header_text,
            font=("TkDefaultFont", 13, "bold"),
            foreground="#1a3a5c",
        ).pack(anchor="w", pady=(0, 4))
        tk.Label(
            self._liq_right_frame,
            text=fecha,
            foreground="#888",
            font=("TkDefaultFont", 9, "italic"),
        ).pack(anchor="w", pady=(0, 12))

        # ─── Lista de links del día ───
        links = local_store.get_links_dia(fecha)
        if not links:
            ttk.Label(
                self._liq_right_frame,
                text="(no hay links para este día)",
                foreground="#888",
                font=("TkDefaultFont", 10, "italic"),
            ).pack(anchor="w", pady=(0, 12))
        else:
            ttk.Label(
                self._liq_right_frame,
                text=f"Links de liquidación ({len(links)}):",
                font=("TkDefaultFont", 10, "bold"),
            ).pack(anchor="w", pady=(0, 6))

            for idx, link in enumerate(links):
                row = ttk.Frame(self._liq_right_frame)
                row.pack(fill="x", pady=2)

                tk.Label(
                    row,
                    text=f"{idx + 1}.",
                    width=3,
                    anchor="ne",
                    foreground="#666",
                ).pack(side="left", padx=(0, 4))

                # Botón ✕ a la derecha
                tk.Button(
                    row,
                    text="✕",
                    command=lambda i=idx: self._liq_remove_link(i),
                    relief="flat",
                    foreground="#c0392b",
                    cursor="hand2",
                    borderwidth=0,
                    font=("TkDefaultFont", 10, "bold"),
                ).pack(side="right")

                # Botón abrir
                tk.Button(
                    row,
                    text="🔗 Abrir",
                    command=lambda u=link: self._open_url(u),
                    cursor="hand2",
                    relief="flat",
                    foreground="#1f4e9d",
                    borderwidth=0,
                    font=("TkDefaultFont", 9, "bold"),
                ).pack(side="right", padx=(0, 4))

                # Texto del link (clickeable también, abre)
                lbl = tk.Label(
                    row,
                    text=link,
                    foreground="#1f4e9d",
                    cursor="hand2",
                    anchor="w",
                    wraplength=320,
                    justify="left",
                )
                lbl.pack(side="left", fill="x", expand=True)
                lbl.bind("<Button-1>", lambda _e, u=link: self._open_url(u))

        # ─── Agregar link nuevo ───
        ttk.Separator(
            self._liq_right_frame, orient="horizontal"
        ).pack(fill="x", pady=(12, 8))

        ttk.Label(
            self._liq_right_frame,
            text="Pegá un link nuevo:",
            font=("TkDefaultFont", 10, "bold"),
        ).pack(anchor="w")

        self._liq_new_link_var = tk.StringVar()
        entry = ttk.Entry(
            self._liq_right_frame,
            textvariable=self._liq_new_link_var,
        )
        entry.pack(fill="x", pady=(2, 6))

        ttk.Button(
            self._liq_right_frame,
            text="+ Agregar link",
            command=self._liq_add_link,
        ).pack(anchor="w")

        entry.bind("<Return>", lambda _e: self._liq_add_link())
        entry.bind("<KP_Enter>", lambda _e: self._liq_add_link())

    def _liq_select_day(self, fecha: str):
        self._liq_selected_date = fecha
        self._liq_render_calendar()
        self._liq_render_right_panel()

    def _liq_prev_month(self):
        if self._liq_month == 1:
            self._liq_month = 12
            self._liq_year -= 1
        else:
            self._liq_month -= 1
        self._liq_render_month_label()
        self._liq_render_calendar()

    def _liq_next_month(self):
        if self._liq_month == 12:
            self._liq_month = 1
            self._liq_year += 1
        else:
            self._liq_month += 1
        self._liq_render_month_label()
        self._liq_render_calendar()

    def _liq_go_today(self):
        today = date.today()
        self._liq_year = today.year
        self._liq_month = today.month
        self._liq_selected_date = today.isoformat()
        self._liq_refresh_all()

    def _liq_add_link(self):
        link = (self._liq_new_link_var.get() or "").strip()
        if not link:
            self._flash_status("Pegá un link primero")
            return
        local_store.add_link_dia(self._liq_selected_date, link)
        self._liq_new_link_var.set("")
        self._liq_refresh_today_banner()
        self._liq_render_calendar()
        self._liq_render_right_panel()
        self._flash_status("Link agregado ✓")

    def _liq_remove_link(self, idx: int):
        local_store.remove_link_dia(self._liq_selected_date, idx)
        self._liq_refresh_today_banner()
        self._liq_render_calendar()
        self._liq_render_right_panel()
        self._flash_status("Link borrado")

    def _on_select(self, _event=None):
        sel = self.tree.selection()
        if not sel:
            return
        leaf_id = sel[0]
        info = self.leaf_to_item.get(leaf_id)
        # Antes de cambiar de selección, persistir la nota de la venta
        # anterior si quedó editada y no la guardó el FocusOut.
        self._flush_nota_pendiente()
        if not info:
            # Es una fila de día, no una venta
            self._update_detail(None, None, None)
            self._load_nota_into_widget(None)
            return
        sku = info.get("sku") or ""
        order_id = self.row_to_order.get(leaf_id)
        self._update_detail(sku, info, order_id)
        self._load_nota_into_widget(order_id)

    def _update_detail(self, sku: str | None, info: dict | None, order_id: str | None):
        # Esconder el panel entero durante el rebuild para que Tk no haga
        # relayout por cada widget destruido/creado (~60-70 widgets).
        self.detail_canvas.itemconfigure(self._detail_window_id, state="hidden")

        # Limpiar contenido anterior
        for f in (self.detail_tags_frame, self.detail_payment_frame,
                  self.detail_costo_frame, self.detail_andres_frame,
                  self.detail_ganancia_frame):
            for w in f.winfo_children():
                w.destroy()

        # Reset de los costos calculados (los setea _render_costo si puede).
        self._last_costo_unitario = None
        self._last_andres_pago_unitario = None
        self._last_pago_per_unit = None
        self._last_cobro_per_unit = None
        self._last_pack_mult = None
        self._render_costo(sku, info)
        self._render_andres(info)
        self._render_payment(info, order_id)
        self._render_ganancia(info, order_id)

        # Mostrar el panel de nuevo.
        self.detail_canvas.itemconfigure(self._detail_window_id, state="normal")

        # Cachear SKU para el picker de etiquetas.
        self._detail_current_sku = sku or None

        if sku is None:
            self.detail_status_var.set("Seleccioná una venta")
            self.detail_status_lbl.configure(foreground="#888")
            self._refresh_tag_picker()
            return

        title = (info or {}).get("title") or ""
        if not sku:
            self.detail_status_var.set(f"(sin SKU)\n{title}" if title else "(sin SKU)")
            self.detail_status_lbl.configure(foreground="#888")
        else:
            self.detail_status_var.set(f"{sku}\n{title}")
            self.detail_status_lbl.configure(foreground="#333")
        self._refresh_tag_picker()

    # ────────────── Etiquetas locales (chips + dropdown) ──────────────

    def _refresh_tag_picker(self):
        """Redibuja chips de etiquetas asignadas + actualiza el combobox con
        las etiquetas del catálogo que todavía no están asignadas a este SKU."""
        for w in self.detail_tags_frame.winfo_children():
            w.destroy()

        sku = self._detail_current_sku
        if not sku:
            ttk.Label(
                self.detail_tags_frame,
                text="(sin SKU para etiquetas)",
                foreground="#888",
            ).pack(anchor="w")
            self._tag_combo.configure(values=[], state="disabled")
            self._tag_combo_var.set("")
            return

        asignadas = local_store.get_etiquetas_sku(sku)
        if not asignadas:
            ttk.Label(
                self.detail_tags_frame,
                text="(sin etiquetas)",
                foreground="#888",
            ).pack(anchor="w")
        else:
            for et in asignadas:
                row = tk.Frame(self.detail_tags_frame, background="#dce6f0")
                row.pack(anchor="w", pady=2)
                tk.Label(
                    row,
                    text=et,
                    background="#dce6f0",
                    foreground="#1a3a5c",
                    padx=8,
                    pady=2,
                    borderwidth=0,
                ).pack(side="left")
                tk.Label(
                    row,
                    text=" ✕ ",
                    background="#dce6f0",
                    foreground="#c0392b",
                    cursor="hand2",
                    padx=2,
                ).pack(side="left")
                # Bind del ✕ — captura el nombre actual.
                row.winfo_children()[1].bind(
                    "<Button-1>",
                    lambda _e, et=et: self._on_remove_tag(et),
                )

        # Combobox: catálogo menos lo ya asignado.
        catalogo = local_store.etiquetas_catalogo()
        disponibles = [e for e in catalogo if e not in asignadas]
        if disponibles:
            self._tag_combo.configure(values=disponibles, state="readonly")
            self._tag_combo_var.set("")
        else:
            self._tag_combo.configure(values=[], state="disabled")
            self._tag_combo_var.set("")

    def _on_add_tag_click(self):
        sku = self._detail_current_sku
        et = (self._tag_combo_var.get() or "").strip()
        if not sku or not et:
            return
        local_store.add_etiqueta_a_sku(sku, et)
        self._refresh_tag_picker()

    def _on_remove_tag(self, etiqueta: str):
        sku = self._detail_current_sku
        if not sku:
            return
        local_store.remove_etiqueta_de_sku(sku, etiqueta)
        self._refresh_tag_picker()

    def _on_new_tag_click(self):
        """Modal chico para crear una etiqueta nueva al catálogo."""
        win = tk.Toplevel(self.root)
        win.title("Nueva etiqueta")
        win.transient(self.root)
        win.resizable(False, False)
        frame = ttk.Frame(win, padding=15)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Nombre de la etiqueta:").pack(anchor="w")
        var = tk.StringVar()
        entry = ttk.Entry(frame, textvariable=var, width=24)
        entry.pack(anchor="w", pady=(2, 12))
        entry.focus_set()

        btns = ttk.Frame(frame)
        btns.pack(fill="x")

        def do_save():
            et = (var.get() or "").strip()
            if not et:
                win.destroy()
                return
            agregada = local_store.add_etiqueta_catalogo(et)
            # Si hay un SKU activo, asignarla automáticamente.
            sku = self._detail_current_sku
            if sku:
                local_store.add_etiqueta_a_sku(sku, et)
            win.destroy()
            self._refresh_tag_picker()
            if agregada:
                self._flash_status(f"Etiqueta agregada: {et}")
            else:
                self._flash_status(f"Etiqueta ya existía: {et}")

        ttk.Button(btns, text="Guardar", command=do_save).pack(side="right")
        ttk.Button(btns, text="Cancelar", command=win.destroy).pack(
            side="right", padx=(0, 6)
        )
        entry.bind("<Return>", lambda _e: do_save())
        win.bind("<Return>", lambda _e: do_save())
        win.bind("<KP_Enter>", lambda _e: do_save())
        win.bind("<Escape>", lambda _e: win.destroy())

        win.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() - win.winfo_width()) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() - win.winfo_height()) // 3
        win.geometry(f"+{x}+{y}")
        win.grab_set()

    def _render_payment(self, info: dict | None, order_id: str | None = None):
        if not info or info.get("payment_id") is None:
            ttk.Label(
                self.detail_payment_frame,
                text="(sin datos de pago)",
                foreground="#888",
            ).pack(anchor="w")
            return

        # Bruto sí lo mostramos como referencia — viene gratis del listado.
        bruto = float(info.get("total_amount") or 0)
        bruto_row = ttk.Frame(self.detail_payment_frame)
        bruto_row.pack(fill="x", anchor="w", pady=1)
        ttk.Label(
            bruto_row, text="Bruto", font=("TkDefaultFont", 10)
        ).pack(side="left")
        tk.Label(
            bruto_row,
            text=format_price(bruto),
            foreground="#1a3a5c",
            font=("TkDefaultFont", 10, "bold"),
        ).pack(side="right")

        # Neto MP + envío manual (ambos cargados a mano).
        neto = local_store.get_neto_manual(order_id) if order_id else None
        shipping = local_store.get_shipping_manual(order_id) if order_id else None

        ttk.Separator(
            self.detail_payment_frame, orient="horizontal"
        ).pack(fill="x", pady=(6, 4))

        if neto is None:
            tk.Label(
                self.detail_payment_frame,
                text="⚠️ Falta neto MP",
                foreground="#c0392b",
                font=("TkDefaultFont", 11, "bold"),
            ).pack(anchor="w")
            ttk.Label(
                self.detail_payment_frame,
                text="Copialo de Mercado Pago y cargalo abajo.",
                foreground="#888",
                font=("TkDefaultFont", 9, "italic"),
            ).pack(anchor="w", pady=(0, 4))
        else:
            # Línea Neto MP (lo que MP te depositó).
            net_row = ttk.Frame(self.detail_payment_frame)
            net_row.pack(fill="x", anchor="w", pady=(0, 2))
            ttk.Label(
                net_row,
                text="Neto MP",
                font=("TkDefaultFont", 10),
            ).pack(side="left")
            tk.Label(
                net_row,
                text=format_price(neto),
                foreground="#1f4e9d",
                font=("TkDefaultFont", 10, "bold"),
            ).pack(side="right")

            # Si hay shipping manual cargado, mostrarlo restando.
            if shipping:
                ship_row = ttk.Frame(self.detail_payment_frame)
                ship_row.pack(fill="x", anchor="w", pady=1)
                ttk.Label(
                    ship_row,
                    text="Envío (Flex)",
                    font=("TkDefaultFont", 10),
                ).pack(side="left")
                tk.Label(
                    ship_row,
                    text=f"- {format_price(shipping)}",
                    foreground="#c0392b",
                    font=("TkDefaultFont", 10, "bold"),
                ).pack(side="right")

            # Total efectivo: si hay envío cargado, neto - envío. Si no, igual al neto.
            ttk.Separator(
                self.detail_payment_frame, orient="horizontal"
            ).pack(fill="x", pady=(4, 2))
            efectivo = neto - (shipping or 0)
            efectivo_row = ttk.Frame(self.detail_payment_frame)
            efectivo_row.pack(fill="x", anchor="w", pady=(0, 2))
            tk.Label(
                efectivo_row,
                text="Neto efectivo →",
                foreground="#1f4e9d",
                font=("TkDefaultFont", 12, "bold"),
            ).pack(side="left")
            tk.Label(
                efectivo_row,
                text=format_price(efectivo),
                foreground="#1f4e9d",
                font=("TkDefaultFont", 12, "bold", "underline"),
            ).pack(side="right")

        # Botones para cargar/editar ambos valores.
        btns_pago = ttk.Frame(self.detail_payment_frame)
        btns_pago.pack(fill="x", anchor="w", pady=(6, 0))
        ttk.Button(
            btns_pago,
            text="✏️ Neto MP",
            command=lambda oid=order_id: self._open_neto_modal(oid),
        ).pack(side="left")
        ttk.Button(
            btns_pago,
            text="🚚 Envío",
            command=lambda oid=order_id: self._open_envio_modal(oid),
        ).pack(side="left", padx=(6, 0))

        method = info.get("payment_method") or ""
        if method:
            ttk.Label(
                self.detail_payment_frame,
                text=f"Método: {method}",
                foreground="#888",
                font=("TkDefaultFont", 9),
            ).pack(anchor="w", pady=(6, 0))

        ttk.Label(
            self.detail_payment_frame,
            text=f"Pago #{info.get('payment_id')}",
            foreground="#888",
            font=("TkDefaultFont", 9),
        ).pack(anchor="w")

    def _open_neto_modal(self, order_id: str | None):
        """Modal chico para cargar/editar el neto MP de una venta."""
        if not order_id:
            self._flash_status("Seleccioná una venta antes de cargar el neto")
            return
        win = tk.Toplevel(self.root)
        win.title("Neto Mercado Pago")
        win.transient(self.root)
        win.resizable(False, False)

        frame = ttk.Frame(win, padding=15)
        frame.pack(fill="both", expand=True)

        ttk.Label(
            frame,
            text=f"Venta #{order_id}",
            font=("TkDefaultFont", 11, "bold"),
        ).pack(anchor="w", pady=(0, 4))
        ttk.Label(
            frame,
            text="Copialo del detalle de la venta en Mercado Pago.",
            foreground="#666",
        ).pack(anchor="w", pady=(0, 12))

        ttk.Label(frame, text="Neto MP (ARS):").pack(anchor="w")
        actual = local_store.get_neto_manual(order_id)
        neto_var = tk.StringVar(value=f"{actual:.2f}" if actual else "")
        entry_row = ttk.Frame(frame)
        entry_row.pack(anchor="w", pady=(2, 12))
        entry = ttk.Entry(entry_row, textvariable=neto_var, width=20)
        entry.pack(side="left")

        def do_format():
            """'$29.750,28' → '29750.28'"""
            raw = neto_var.get().strip()
            # Sacar todo menos dígitos, puntos y comas
            cleaned = ""
            for ch in raw:
                if ch.isdigit() or ch in ".,":
                    cleaned += ch
            if not cleaned:
                return
            # Detectar formato AR: si hay coma seguida de 1-2 dígitos al final,
            # es el separador decimal; los puntos son miles.
            m = re.match(r'^([\d.]+),(\d{1,2})$', cleaned)
            if m:
                integer_part = m.group(1).replace(".", "")
                decimal_part = m.group(2)
                cleaned = f"{integer_part}.{decimal_part}"
            else:
                # Sin coma: los puntos podrían ser miles (ej "29.750")
                # o decimal (ej "29750.28"). Si hay un punto seguido de
                # exactamente 3 dígitos al final, es separador de miles.
                m2 = re.match(r'^([\d.]+)\.(\d{3})$', cleaned)
                if m2:
                    cleaned = cleaned.replace(".", "")
            try:
                val = float(cleaned)
                neto_var.set(f"{val:.2f}")
                entry.select_range(0, "end")
            except ValueError:
                pass

        ttk.Button(entry_row, text="Limpiar", width=8, command=do_format).pack(
            side="left", padx=(6, 0)
        )
        entry.focus_set()
        entry.select_range(0, "end")

        ttk.Label(
            frame,
            text="Dejar vacío o 0 para borrar el neto cargado.",
            foreground="#888",
            font=("TkDefaultFont", 9, "italic"),
        ).pack(anchor="w", pady=(0, 8))

        btns = ttk.Frame(frame)
        btns.pack(fill="x")

        def do_save():
            # Punto como separador decimal. Acepta también coma por las dudas
            # (la convertimos a punto). NO interpretamos punto como miles.
            raw = neto_var.get().strip().replace(",", ".")
            if raw == "":
                new_neto = None
            else:
                try:
                    new_neto = float(raw)
                except ValueError:
                    self._flash_status("Valor inválido — debe ser un número")
                    return
                if new_neto < 0:
                    self._flash_status("El neto no puede ser negativo")
                    return
            try:
                local_store.set_neto_manual(order_id, new_neto)
            except Exception as e:
                messagebox.showerror(
                    "Error", f"No se pudo guardar el neto:\n{e}", parent=win
                )
                return
            win.destroy()
            self._flash_status(f"Neto MP guardado para venta {order_id} ✓")
            self._on_select()
            self._update_totales_inline()

        ttk.Button(btns, text="Guardar", command=do_save).pack(side="right")
        ttk.Button(btns, text="Cancelar", command=win.destroy).pack(
            side="right", padx=(0, 6)
        )
        entry.bind("<Return>", lambda _e: do_save())
        win.bind("<Return>", lambda _e: do_save())
        win.bind("<KP_Enter>", lambda _e: do_save())
        win.bind("<Escape>", lambda _e: win.destroy())

        win.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() - win.winfo_width()) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() - win.winfo_height()) // 3
        win.geometry(f"+{x}+{y}")
        win.grab_set()

    def _open_envio_modal(self, order_id: str | None):
        """Modal para cargar/editar el costo de envío (Flex) de una venta.
        Si está cargado, se resta del neto MP en todos los cálculos."""
        if not order_id:
            self._flash_status("Seleccioná una venta antes de cargar el envío")
            return
        win = tk.Toplevel(self.root)
        win.title("Costo de envío")
        win.transient(self.root)
        win.resizable(False, False)

        frame = ttk.Frame(win, padding=15)
        frame.pack(fill="both", expand=True)

        ttk.Label(
            frame,
            text=f"Venta #{order_id}",
            font=("TkDefaultFont", 11, "bold"),
        ).pack(anchor="w", pady=(0, 4))
        ttk.Label(
            frame,
            text="Lo que vas a pagar afuera por la entrega (Flex/Héctor).",
            foreground="#666",
        ).pack(anchor="w", pady=(0, 12))

        ttk.Label(frame, text="Costo envío (ARS):").pack(anchor="w")
        actual = local_store.get_shipping_manual(order_id)
        env_var = tk.StringVar(value=f"{actual:.2f}" if actual else "")
        entry = ttk.Entry(frame, textvariable=env_var, width=20)
        entry.pack(anchor="w", pady=(2, 12))
        entry.focus_set()
        entry.select_range(0, "end")

        ttk.Label(
            frame,
            text="Dejar vacío o 0 para borrar el envío cargado.",
            foreground="#888",
            font=("TkDefaultFont", 9, "italic"),
        ).pack(anchor="w", pady=(0, 8))

        btns = ttk.Frame(frame)
        btns.pack(fill="x")

        def do_save():
            # Punto como separador decimal. Acepta también coma por las dudas.
            raw = env_var.get().strip().replace(",", ".")
            if raw == "":
                new_env = None
            else:
                try:
                    new_env = float(raw)
                except ValueError:
                    self._flash_status("Valor inválido — debe ser un número")
                    return
                if new_env < 0:
                    self._flash_status("El envío no puede ser negativo")
                    return
            try:
                local_store.set_shipping_manual(order_id, new_env)
            except Exception as e:
                messagebox.showerror(
                    "Error", f"No se pudo guardar el envío:\n{e}", parent=win
                )
                return
            win.destroy()
            self._flash_status(f"Envío guardado para venta {order_id} ✓")
            self._on_select()
            self._update_totales_inline()

        ttk.Button(btns, text="Guardar", command=do_save).pack(side="right")
        ttk.Button(btns, text="Cancelar", command=win.destroy).pack(
            side="right", padx=(0, 6)
        )
        entry.bind("<Return>", lambda _e: do_save())
        win.bind("<Return>", lambda _e: do_save())
        win.bind("<KP_Enter>", lambda _e: do_save())
        win.bind("<Escape>", lambda _e: win.destroy())

        win.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() - win.winfo_width()) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() - win.winfo_height()) // 3
        win.geometry(f"+{x}+{y}")
        win.grab_set()

    def _render_ganancia(self, info: dict | None, order_id: str | None = None):
        frame = self.detail_ganancia_frame
        costo_unitario = self._last_costo_unitario
        # neto efectivo = neto MP - envío manual (si está cargado).
        neto = local_store.get_neto_efectivo(order_id) if order_id else None

        if costo_unitario is None or neto is None:
            faltan = []
            if costo_unitario is None:
                faltan.append("FOB/multiplicador")
            if neto is None:
                faltan.append("neto MP")
            ttk.Label(
                frame,
                text=f"(falta {' + '.join(faltan)})" if faltan else "(faltan datos)",
                foreground="#888",
            ).pack(anchor="w")
            return

        try:
            quantity = int((info or {}).get("quantity") or 1)
        except (TypeError, ValueError):
            quantity = 1

        costo_total = costo_unitario * quantity
        ganancia = float(neto) - costo_total
        color = "#1e7a1e" if ganancia >= 0 else "#c0392b"

        # Desglose chico arriba para que se entienda de dónde sale.
        rows_detail = [
            ("Neto MP", format_price(neto), "#1e7a1e"),
            (
                f"Costo (×{quantity})" if quantity != 1 else "Costo",
                f"- {format_price(costo_total)}",
                "#c0392b",
            ),
        ]
        for label, value, value_color in rows_detail:
            row = ttk.Frame(frame)
            row.pack(fill="x", anchor="w", pady=1)
            ttk.Label(row, text=label, font=("TkDefaultFont", 10)).pack(side="left")
            tk.Label(
                row,
                text=value,
                foreground=value_color,
                font=("TkDefaultFont", 10),
            ).pack(side="right")

        # Línea final destacada.
        row = ttk.Frame(frame)
        row.pack(fill="x", anchor="w", pady=(6, 0))
        tk.Label(
            row,
            text="Ganancia Total Pablo →",
            foreground="#1e7a1e",
            font=("TkDefaultFont", 12, "bold"),
        ).pack(side="left")
        tk.Label(
            row,
            text=format_price(ganancia),
            foreground="#1e7a1e",
            font=("TkDefaultFont", 12, "bold", "underline"),
        ).pack(side="right")

        # Margen sobre el bruto (precio de venta listado en ML).
        # Es la métrica estándar de "comercio": de cada peso vendido,
        # cuántos centavos quedan como ganancia neta para Pablo.
        # Thresholds aproximados para revendedores de importación en ML:
        #   < 10%   → MUY BAJO  (rojo)    no rinde después del laburo
        #   10-20%  → BAJO      (naranja) rentable pero ajustado
        #   20-30%  → BUENO     (verde)   margen sano de importador
        #   > 30%   → EXCELENTE (violeta) producto ganador
        def chip_for(pct):
            if pct < 10:
                return "MUY BAJO", "#c0392b"
            elif pct < 20:
                return "BAJO", "#d35400"
            elif pct < 30:
                return "BUENO", "#1e7a1e"
            else:
                return "EXCELENTE", "#7d3c98"

        total_amount = float((info or {}).get("total_amount") or 0)
        if total_amount > 0:
            margen_pct = (ganancia / total_amount) * 100
            chip_text, chip_bg = chip_for(margen_pct)

            margen_row = ttk.Frame(frame)
            margen_row.pack(fill="x", anchor="w", pady=(8, 0))
            ttk.Label(
                margen_row,
                text="Margen Pablo",
                font=("TkDefaultFont", 11, "bold"),
            ).pack(side="left")
            tk.Label(
                margen_row,
                text=f"{margen_pct:.1f}%",
                foreground="#1a3a5c",
                font=("TkDefaultFont", 12, "bold"),
            ).pack(side="left", padx=(8, 8))
            tk.Label(
                margen_row,
                text=f"  {chip_text}  ",
                foreground="white",
                background=chip_bg,
                font=("TkDefaultFont", 9, "bold"),
            ).pack(side="left")

            # Leyenda de niveles (italic gris) — solo se imprime una vez,
            # debajo del margen de Pablo, porque Andrés usa los mismos.
            ttk.Label(
                frame,
                text="<10 MUY BAJO  ·  10-20 BAJO  ·  20-30 BUENO  ·  >30 EXCELENTE",
                foreground="#888",
                font=("TkDefaultFont", 9, "italic"),
            ).pack(anchor="w", pady=(1, 0))

            # Sugerencia de precio para llegar al próximo tier (BUENO/EXCELENTE).
            # Asume que la relación neto/bruto se mantiene constante cuando se
            # cambia el precio (los fees de MP escalan ~lineal). Es una
            # aproximación: el envío manual fijo introduce algo de error pero
            # alcanza para una sugerencia útil.
            def precio_pablo_para(target_pct):
                ratio = neto / total_amount
                denom = ratio - target_pct / 100
                if denom <= 0:
                    return None
                return costo_total / denom

            partes_p = []
            if margen_pct < 20:
                p = precio_pablo_para(20)
                if p is not None:
                    partes_p.append(f"BUENO vendé a {format_price(p)}")
            if margen_pct < 30:
                p = precio_pablo_para(30)
                if p is not None:
                    partes_p.append(f"EXCELENTE vendé a {format_price(p)}")
            if partes_p:
                ttk.Label(
                    frame,
                    text="Para " + "  ·  Para ".join(partes_p),
                    foreground="#888",
                    font=("TkDefaultFont", 9, "italic"),
                ).pack(anchor="w")

        # Ganancia total de Andrés primero, luego margen con chip.
        cobro_u = self._last_cobro_per_unit
        pago_u = self._last_pago_per_unit
        pack_mult = self._last_pack_mult or 1
        if cobro_u is not None and pago_u is not None:
            try:
                qty = int((info or {}).get("quantity") or 1)
            except (TypeError, ValueError):
                qty = 1
            gan_andres_total = (cobro_u - pago_u) * pack_mult * qty
            row_ga = ttk.Frame(frame)
            row_ga.pack(fill="x", anchor="w", pady=(8, 0))
            tk.Label(
                row_ga,
                text="Ganancia Total Andrés →",
                foreground="#7d3c98",
                font=("TkDefaultFont", 12, "bold"),
            ).pack(side="left")
            tk.Label(
                row_ga,
                text=format_price(gan_andres_total),
                foreground="#7d3c98",
                font=("TkDefaultFont", 12, "bold", "underline"),
            ).pack(side="right")

        # Margen de Andrés sobre su costo (= markup − 1).
        sku = (info or {}).get("sku") if info else None
        markup_a = local_store.get_markup(sku) if sku else None
        if markup_a is None:
            markup_a = GANANCIA_HERMANO_MULT
        margen_a_pct = (markup_a - 1) * 100
        chip_text_a, chip_bg_a = chip_for(margen_a_pct)

        margen_a_row = ttk.Frame(frame)
        margen_a_row.pack(fill="x", anchor="w", pady=(6, 0))
        ttk.Label(
            margen_a_row,
            text="Margen Andrés",
            font=("TkDefaultFont", 11, "bold"),
        ).pack(side="left")
        tk.Label(
            margen_a_row,
            text=f"{margen_a_pct:.1f}%",
            foreground="#1a3a5c",
            font=("TkDefaultFont", 12, "bold"),
        ).pack(side="left", padx=(8, 8))
        tk.Label(
            margen_a_row,
            text=f"  {chip_text_a}  ",
            foreground="white",
            background=chip_bg_a,
            font=("TkDefaultFont", 9, "bold"),
        ).pack(side="left")

        # Sugerencia de markup para Andrés. Trivial: target_pct% → 1 + t/100.
        partes_a = []
        if margen_a_pct < 20:
            partes_a.append("BUENO subí markup a 1.20")
        if margen_a_pct < 30:
            partes_a.append("EXCELENTE subí markup a 1.30")
        if partes_a:
            ttk.Label(
                frame,
                text="Para " + "  ·  Para ".join(partes_a),
                foreground="#888",
                font=("TkDefaultFont", 9, "italic"),
            ).pack(anchor="w", pady=(1, 0))

    # ────────────── Notas por venta ──────────────

    def _load_nota_into_widget(self, order_id: str | None):
        """Carga la nota persistida en el Text widget. Marca _nota_loading
        para que el FocusOut no la "guarde" como input del usuario."""
        self._nota_loading = True
        self._nota_current_order = order_id
        self.detail_nota_text.delete("1.0", "end")
        if order_id:
            nota = local_store.get_nota(order_id)
            if nota:
                self.detail_nota_text.insert("1.0", nota)
        self._nota_loading = False

    def _on_nota_focus_out(self, _event=None):
        """Cuando el usuario sale del Text widget, persiste si cambió."""
        if self._nota_loading:
            return
        order_id = self._nota_current_order
        if not order_id:
            return
        nuevo = self.detail_nota_text.get("1.0", "end-1c").strip()
        viejo = local_store.get_nota(order_id)
        if nuevo == viejo:
            return
        local_store.set_nota(order_id, nuevo)
        self._refresh_leaf_nota_tag(order_id)
        self._update_status()

    def _flush_nota_pendiente(self):
        """Llamado al cambiar de selección: si la nota actual está editada
        y no se guardó (porque el usuario no perdió foco todavía), guardala
        ahora antes de que se sobreescriba con la nota de la venta nueva."""
        if self._nota_loading:
            return
        order_id = self._nota_current_order
        if not order_id:
            return
        nuevo = self.detail_nota_text.get("1.0", "end-1c").strip()
        viejo = local_store.get_nota(order_id)
        if nuevo == viejo:
            return
        local_store.set_nota(order_id, nuevo)
        self._refresh_leaf_nota_tag(order_id)

    def _refresh_leaf_nota_tag(self, order_id: str):
        """Actualiza el tag visual del row del tree para que aparezca/desaparezca
        el foreground violeta de "tiene nota"."""
        for leaf_id, oid in self.row_to_order.items():
            if oid != order_id:
                continue
            try:
                self.tree.item(
                    leaf_id, tags=self._row_tags(leaf_id, order_id)
                )
            except tk.TclError:
                pass
        # Si el filtro "solo con nota" está activo, refrescar el tree
        # porque la fila puede haber entrado o salido del set visible.
        if self._solo_con_nota_var.get():
            self._refresh_tree_filter()

    def _render_costo(self, sku: str | None, info: dict | None = None):
        frame = self.detail_costo_frame
        if not sku:
            ttk.Label(
                frame, text="Sin SKU — cargá el SKU para costear", foreground="#888",
                font=("TkDefaultFont", 9, "italic"),
            ).pack(anchor="w")
            return
        fob = local_store.get_fob(sku)
        if not fob or fob <= 0:
            ttk.Label(
                frame,
                text="Sin precio FOB cargado",
                foreground="#888",
            ).pack(anchor="w")
            ttk.Label(
                frame,
                text="Ctrl+Click en la fila para cargarlo",
                foreground="#888",
                font=("TkDefaultFont", 9, "italic"),
            ).pack(anchor="w")
            return

        mult = local_store.get_multiplicador(sku)
        if mult is None:
            # FOB cargado pero falta el multiplicador → no se puede calcular nada.
            row = ttk.Frame(frame)
            row.pack(fill="x", anchor="w", pady=1)
            ttk.Label(row, text="FOB unitario", font=("TkDefaultFont", 10)).pack(side="left")
            tk.Label(
                row,
                text=f"USD {fob:,.2f}",
                foreground="#d35400",
                font=("TkDefaultFont", 10, "bold"),
            ).pack(side="right")
            ttk.Label(
                frame,
                text="⚠️ Falta multiplicador",
                foreground="#d35400",
                font=("TkDefaultFont", 10, "bold"),
            ).pack(anchor="w", pady=(6, 0))
            ttk.Label(
                frame,
                text="Alt+Click en la fila para cargarlo",
                foreground="#888",
                font=("TkDefaultFont", 9, "italic"),
            ).pack(anchor="w")
            return

        fob_total = fob * mult
        nacionalizado_usd = fob_total * NACIONALIZACION_MULT
        cot = dolar.get()
        if cot is None:
            if not dolar.loaded():
                ttk.Label(
                    frame, text="Cargando dólar…", foreground="#888"
                ).pack(anchor="w")
            else:
                ttk.Label(
                    frame,
                    text=f"FOB total: USD {fob_total:,.2f}",
                    foreground="#1a3a5c",
                ).pack(anchor="w")
                ttk.Label(
                    frame,
                    text="⚠️ No se pudo obtener el dólar",
                    foreground="#c0392b",
                ).pack(anchor="w")
            return

        # ── Cálculo paso a paso, todo por unidad PRIMERO ──
        # Construimos los valores per-unit (1 sola unidad física), después
        # multiplicamos por el pack y al final por la cantidad de la venta.
        # Markup de Andrés: por defecto GANANCIA_HERMANO_MULT, pero se puede
        # override por SKU (persistido en data.json).
        markup = local_store.get_markup(sku)
        if markup is None:
            markup = GANANCIA_HERMANO_MULT
        nac_usd_unit = fob * NACIONALIZACION_MULT          # USD por unidad
        pesos_unit = nac_usd_unit * cot                     # $ por unidad
        andres_unit = pesos_unit * markup                   # $ por unidad con markup

        # Por pack (= por una venta de 1 cantidad). El multiplicador del pack
        # se aplica acá, no antes.
        andres_pack = andres_unit * mult
        pago_pack = pesos_unit * mult  # lo que Andrés pagó en China por pack

        # Cantidad de la venta (cuántos packs se vendieron en esta orden).
        try:
            quantity = int((info or {}).get("quantity") or 1)
        except (TypeError, ValueError):
            quantity = 1
        andres_total = andres_pack * quantity

        # Variables para el resto de la app. Mantenemos la semántica per-pack
        # porque _render_ganancia y _calcular_totales_seleccionados ya esperan
        # eso (después multiplican por quantity).
        self._last_costo_unitario = andres_pack
        self._last_andres_pago_unitario = pago_pack
        # Valores per-unidad real (1 sola unidad física) + mult del pack para
        # que _render_andres pueda armar el desglose paso a paso (individual →
        # pack → total venta) en sintonía con esta sección.
        self._last_pago_per_unit = pesos_unit
        self._last_cobro_per_unit = andres_unit
        self._last_pack_mult = mult

        # ── Render: paso a paso, per-unidad ──
        rows_unit = [
            ("FOB unitario", f"USD {fob:,.2f}", "#1a3a5c"),
            (f"Nacionalizado (×{NACIONALIZACION_MULT})",
             f"USD {nac_usd_unit:,.2f}", "#1a3a5c"),
        ]
        for label, value, color in rows_unit:
            row = ttk.Frame(frame)
            row.pack(fill="x", anchor="w", pady=1)
            ttk.Label(row, text=label, font=("TkDefaultFont", 10)).pack(side="left")
            tk.Label(
                row, text=value, foreground=color, font=("TkDefaultFont", 10)
            ).pack(side="right")

        # En pesos: misma estructura pero con chip "Costo individual" al lado
        # del label, porque es el costo real per-unidad sin markup todavía.
        row_pesos = ttk.Frame(frame)
        row_pesos.pack(fill="x", anchor="w", pady=1)
        ttk.Label(
            row_pesos,
            text=f"En pesos (×${cot:,.0f})",
            font=("TkDefaultFont", 10),
        ).pack(side="left")
        tk.Label(
            row_pesos,
            text="  Costo individual  ",
            foreground="white",
            background="#1a3a5c",
            font=("TkDefaultFont", 9, "bold"),
        ).pack(side="left", padx=(6, 0))
        tk.Label(
            row_pesos,
            text=format_price(pesos_unit),
            foreground="#1a3a5c",
            font=("TkDefaultFont", 10),
        ).pack(side="right")

        # Misma línea pero × multiplicador del pack, solo si mult > 1.
        # (Si es por unidad, sería duplicar la fila de arriba.)
        if mult > 1:
            row_pack_pesos = ttk.Frame(frame)
            row_pack_pesos.pack(fill="x", anchor="w", pady=1)
            ttk.Label(
                row_pack_pesos,
                text=f"En pesos (×${cot:,.0f})",
                font=("TkDefaultFont", 10),
            ).pack(side="left")
            tk.Label(
                row_pack_pesos,
                text="  Costo x Pack  ",
                foreground="white",
                background="#1a3a5c",
                font=("TkDefaultFont", 9, "bold"),
            ).pack(side="left", padx=(6, 0))
            tk.Label(
                row_pack_pesos,
                text=f"  Pack {mult}  ",
                foreground="white",
                background="#7d3c98",
                font=("TkDefaultFont", 9, "bold"),
            ).pack(side="left", padx=(4, 0))
            tk.Label(
                row_pack_pesos,
                text=format_price(pesos_unit * mult),
                foreground="#1a3a5c",
                font=("TkDefaultFont", 10),
            ).pack(side="right")

        # Costo total de la venta = costo per pack × cantidad. Solo si > 1.
        try:
            qty_for_total = int((info or {}).get("quantity") or 1)
        except (TypeError, ValueError):
            qty_for_total = 1
        if qty_for_total > 1:
            row_total_pesos = ttk.Frame(frame)
            row_total_pesos.pack(fill="x", anchor="w", pady=1)
            ttk.Label(
                row_total_pesos,
                text=f"En pesos (×${cot:,.0f})",
                font=("TkDefaultFont", 10),
            ).pack(side="left")
            tk.Label(
                row_total_pesos,
                text="  Costo total  ",
                foreground="white",
                background="#1a3a5c",
                font=("TkDefaultFont", 9, "bold"),
            ).pack(side="left", padx=(6, 0))
            tk.Label(
                row_total_pesos,
                text=f"  {qty_for_total} items  ",
                foreground="white",
                background="#7d3c98",
                font=("TkDefaultFont", 9, "bold"),
            ).pack(side="left", padx=(4, 0))
            tk.Label(
                row_total_pesos,
                text=format_price(pesos_unit * mult * qty_for_total),
                foreground="#1a3a5c",
                font=("TkDefaultFont", 10),
            ).pack(side="right")

        # Total individual (1 unidad con markup de Andrés).
        # La línea entera es clickeable → abre modal para editar el markup.
        # Mostramos el markup sin ceros innecesarios (1.3 no 1.30, 1.4 no 1.40).
        markup_str = f"{markup:g}"
        row_ind_bg = "#f0e6f6"
        row_ind = tk.Frame(
            frame, cursor="hand2", background=row_ind_bg,
            highlightbackground="#7d3c98", highlightthickness=2,
            padx=6, pady=4,
        )
        row_ind.pack(fill="x", anchor="w", pady=(4, 1))
        lbl_ind = tk.Label(
            row_ind,
            text="✏️ Pagar a Andrés individual:",
            foreground="#7d3c98", background=row_ind_bg,
            font=("TkDefaultFont", 12, "bold"),
            cursor="hand2",
        )
        lbl_ind.pack(side="left")
        chip_markup = tk.Label(
            row_ind,
            text=f"  $$$ {markup_str} %  ",
            foreground="white",
            background="#d35400",
            font=("TkDefaultFont", 9, "bold"),
            cursor="hand2",
        )
        chip_markup.pack(side="left", padx=(6, 0))
        val_ind = tk.Label(
            row_ind,
            text=format_price(andres_unit),
            foreground="#7d3c98", background=row_ind_bg,
            font=("TkDefaultFont", 12, "bold", "underline"),
            cursor="hand2",
        )
        val_ind.pack(side="right")
        for w in (row_ind, lbl_ind, chip_markup, val_ind):
            w.bind(
                "<Button-1>",
                lambda _e, s=sku, t=(info or {}).get("title") or "": self._open_markup_modal(s, t),
            )

        # ── × multiplicador del pack (solo si > 1) ──
        if mult > 1:
            row_pack = ttk.Frame(frame)
            row_pack.pack(fill="x", anchor="w", pady=(4, 1))
            tk.Label(
                row_pack,
                text="Pagar a Andrés por pack:",
                foreground="#7d3c98",
                font=("TkDefaultFont", 12, "bold"),
            ).pack(side="left")
            tk.Label(
                row_pack,
                text=f"  Pack {mult}  ",
                foreground="white",
                background="#d35400",
                font=("TkDefaultFont", 9, "bold"),
            ).pack(side="left", padx=(6, 0))
            tk.Label(
                row_pack,
                text=format_price(andres_pack),
                foreground="#7d3c98",
                font=("TkDefaultFont", 12, "bold", "underline"),
            ).pack(side="right")

        # ── × cantidad de la venta (solo si > 1) ──
        if quantity > 1:
            row_tot = ttk.Frame(frame)
            row_tot.pack(fill="x", anchor="w", pady=(4, 1))
            tk.Label(
                row_tot,
                text="Pagar a Andrés total venta:",
                foreground="#c0392b",
                font=("TkDefaultFont", 12, "bold"),
            ).pack(side="left")
            tk.Label(
                row_tot,
                text=f"  {quantity} items  ",
                foreground="white",
                background="#d35400",
                font=("TkDefaultFont", 9, "bold"),
            ).pack(side="left", padx=(6, 0))
            tk.Label(
                row_tot,
                text=format_price(andres_total),
                foreground="#c0392b",
                font=("TkDefaultFont", 12, "bold", "underline"),
            ).pack(side="right")

    def _render_andres(self, info: dict | None):
        """Ganancia que se queda Andrés: cobra − paga, con desglose paso a
        paso (individual → pack → total venta), reusando los valores per-
        unidad ya calculados en _render_costo.

          Ganancia Andrés individual:  [$$$ X %]    $X   ← (markup−1) × pesos_unit
          Ganancia Andrés por Pack:    [Pack N]     $X   ← × mult del pack (si > 1)
          Ganancia Andrés total:       [N items]    $X   ← × cantidad venta (si > 1)
        """
        frame = self.detail_andres_frame
        pago_u = self._last_pago_per_unit
        cobro_u = self._last_cobro_per_unit
        mult = self._last_pack_mult or 1
        if pago_u is None or cobro_u is None:
            ttk.Label(
                frame,
                text="(faltan datos para calcular)",
                foreground="#888",
            ).pack(anchor="w")
            return

        try:
            quantity = int((info or {}).get("quantity") or 1)
        except (TypeError, ValueError):
            quantity = 1

        # Markup específico del SKU para mostrar en el chip "$$$ X %".
        sku = (info or {}).get("sku") if info else None
        markup = local_store.get_markup(sku) if sku else None
        if markup is None:
            markup = GANANCIA_HERMANO_MULT
        markup_str = f"{markup:g}"

        gan_unit = cobro_u - pago_u            # ganancia por unidad
        gan_pack = gan_unit * mult             # × mult del pack
        gan_total = gan_pack * quantity        # × cantidad de la venta

        # Paleta nueva para esta sección: chips en teal, líneas en verde
        # oscuro y la total en rojo (visual hierarchy: el "número grande").
        chip_bg = "#16a085"
        line_color_normal = "#117a65"
        line_color_total = "#c0392b"

        def add_row(label_text: str, chip_text: str, value: float, color: str):
            row = ttk.Frame(frame)
            row.pack(fill="x", anchor="w", pady=(4, 1))
            tk.Label(
                row,
                text=label_text,
                foreground=color,
                font=("TkDefaultFont", 12, "bold"),
            ).pack(side="left")
            tk.Label(
                row,
                text=f"  {chip_text}  ",
                foreground="white",
                background=chip_bg,
                font=("TkDefaultFont", 9, "bold"),
            ).pack(side="left", padx=(6, 0))
            tk.Label(
                row,
                text=format_price(value),
                foreground=color,
                font=("TkDefaultFont", 12, "bold", "underline"),
            ).pack(side="right")

        add_row(
            "Ganancia Andrés individual:",
            f"$$$ {markup_str} %",
            gan_unit,
            line_color_normal,
        )
        if mult > 1:
            add_row(
                "Ganancia Andrés por Pack:",
                f"Pack {mult}",
                gan_pack,
                line_color_normal,
            )
        if quantity > 1:
            add_row(
                "Ganancia Andrés total:",
                f"{quantity} items",
                gan_total,
                line_color_total,
            )

    # ────────────────── Atajos de teclado (F1-F6) ──────────────────
    # Estructura: pequeños URL builders + un helper que devuelve la fila
    # seleccionada validada, y cada handler de F-key chequea la selección
    # una sola vez y dispara el browser. F6 reusa los URL builders para
    # abrir las cuatro pestañas sin repetir 4 veces el mensaje de error.

    def _selected_leaf_data(self):
        """(leaf_id, info, order_id) del leaf seleccionado o None si no hay
        venta válida (vacío, día, o info faltante)."""
        sel = self.tree.selection()
        if not sel:
            return None
        leaf_id = sel[0]
        if self.tree.parent(leaf_id) == "":
            return None  # fila de día, no de venta
        info = self.leaf_to_item.get(leaf_id)
        if not info:
            return None
        order_id = self.row_to_order.get(leaf_id)
        return leaf_id, info, order_id

    def _open_detalle_venta(self, order_id: str):
        self._open_url(
            f"https://www.mercadolibre.com.ar/ventas/{order_id}/detalle"
        )

    def _open_pago_mp(self, query_id):
        # /activities?q=ID porque /activities/detail/{id} requiere un hash
        # purchase_v3-{...} impredecible que solo conoce el frontend de MP.
        # query_id puede ser payment_id (preferido) u order_id como fallback.
        self._open_url(
            f"https://www.mercadopago.com.ar/activities?q={query_id}"
        )

    def _open_publi_edit(self, item_id: str):
        # ML redirige /publicaciones/{id}/modificar a la URL larga con el
        # token de sesión, así que la forma corta es estable.
        self._open_url(
            f"https://www.mercadolibre.com.ar/publicaciones/{item_id}/modificar"
        )

    def _open_publi_publica(self, item_id: str):
        # MLA1234567890 → articulo.mercadolibre.com.ar/MLA-1234567890.
        if len(item_id) > 3 and item_id[:3].isalpha():
            url = f"https://articulo.mercadolibre.com.ar/{item_id[:3]}-{item_id[3:]}"
        else:
            url = f"https://articulo.mercadolibre.com.ar/{item_id}"
        self._open_url(url)

    def _on_f1_detalle_venta(self, _e=None):
        sel = self._selected_leaf_data()
        if not sel:
            self._flash_status("Seleccioná una venta primero")
            return "break"
        _, _, order_id = sel
        if not order_id:
            self._flash_status("Esta fila no tiene order_id")
            return "break"
        self._open_detalle_venta(order_id)
        self._flash_status(f"Abriendo venta {order_id}")
        return "break"

    def _on_f2_pago_mp(self, _e=None):
        sel = self._selected_leaf_data()
        if not sel:
            self._flash_status("Seleccioná una venta primero")
            return "break"
        _, info, order_id = sel
        qid = info.get("payment_id") or order_id
        if not qid:
            self._flash_status("No hay payment_id ni order_id")
            return "break"
        self._open_pago_mp(qid)
        self._flash_status(f"Abriendo pago {qid}")
        return "break"

    def _on_f3_edit_publicacion(self, _e=None):
        sel = self._selected_leaf_data()
        if not sel:
            self._flash_status("Seleccioná una venta primero")
            return "break"
        _, info, _ = sel
        item_id = info.get("item_id") or ""
        if not item_id:
            self._flash_status("Esta fila no tiene ID de publicación")
            return "break"
        self._open_publi_edit(item_id)
        self._flash_status(f"Abriendo editor de {item_id}")
        return "break"

    def _on_f4_publi_publica(self, _e=None):
        sel = self._selected_leaf_data()
        if not sel:
            self._flash_status("Seleccioná una venta primero")
            return "break"
        _, info, _ = sel
        item_id = info.get("item_id") or ""
        if not item_id:
            self._flash_status("Esta fila no tiene ID de publicación")
            return "break"
        self._open_publi_publica(item_id)
        self._flash_status(f"Abriendo publicación {item_id}")
        return "break"

    def _on_f6_abrir_todo(self, _e=None):
        """Abre F1+F2+F3+F4 en pestañas separadas. Mensaje de error único."""
        sel = self._selected_leaf_data()
        if not sel:
            self._flash_status("Seleccioná una venta primero")
            return "break"
        _, info, order_id = sel
        item_id = info.get("item_id") or ""
        payment_id = info.get("payment_id") or order_id
        opened = 0
        if order_id:
            self._open_detalle_venta(order_id)
            opened += 1
        if payment_id:
            self._open_pago_mp(payment_id)
            opened += 1
        if item_id:
            self._open_publi_edit(item_id)
            opened += 1
            self._open_publi_publica(item_id)
            opened += 1
        self._flash_status(f"Abriendo {opened} pestañas en el browser")
        return "break"

    # ────────────── Mouse: Ctrl+click FOB / Alt+click multiplicador ──────────────

    def _on_ctrl_click_fob(self, event):
        row = self.tree.identify_row(event.y)
        if not row or self.tree.parent(row) == "":
            return "break"
        info = self.leaf_to_item.get(row)
        if not info:
            return "break"
        sku = info.get("sku") or ""
        if not sku:
            self._flash_status("Esta venta no tiene SKU — no se puede cargar FOB")
            return "break"
        self._open_fob_modal(sku, info.get("title") or "")
        return "break"

    def _open_fob_modal(self, sku: str, title: str):
        win = tk.Toplevel(self.root)
        win.title("Editar precio FOB")
        win.transient(self.root)
        win.resizable(False, False)

        frame = ttk.Frame(win, padding=15)
        frame.pack(fill="both", expand=True)

        ttk.Label(
            frame,
            text=title,
            wraplength=420,
            font=("TkDefaultFont", 11, "bold"),
        ).pack(anchor="w", pady=(0, 4))
        ttk.Label(frame, text=f"SKU: {sku}", foreground="#666").pack(
            anchor="w", pady=(0, 12)
        )

        ttk.Label(frame, text="Precio FOB (USD):").pack(anchor="w")
        actual = local_store.get_fob(sku)
        fob_var = tk.StringVar(value=f"{actual:.2f}" if actual else "")
        entry = ttk.Entry(frame, textvariable=fob_var, width=20)
        entry.pack(anchor="w", pady=(2, 12))
        entry.focus_set()
        entry.select_range(0, "end")

        hint = ttk.Label(
            frame,
            text="Dejar vacío o 0 para borrar el FOB cargado.",
            foreground="#888",
            font=("TkDefaultFont", 9, "italic"),
        )
        hint.pack(anchor="w", pady=(0, 8))

        btns = ttk.Frame(frame)
        btns.pack(fill="x")

        save_btn = ttk.Button(btns, text="Guardar")
        cancel_btn = ttk.Button(btns, text="Cancelar", command=win.destroy)
        save_btn.pack(side="right")
        cancel_btn.pack(side="right", padx=(0, 6))

        def do_save():
            raw = fob_var.get().strip().replace(",", ".")
            if raw == "":
                new_fob = 0.0
            else:
                try:
                    new_fob = float(raw)
                except ValueError:
                    self._flash_status("Valor inválido — debe ser un número")
                    return
                if new_fob < 0:
                    self._flash_status("El FOB no puede ser negativo")
                    return
            try:
                local_store.set_fob(sku, new_fob)
            except Exception as e:
                messagebox.showerror(
                    "Error", f"No se pudo guardar el FOB:\n{e}", parent=win
                )
                return
            self._on_fob_saved(win, sku)

        save_btn.configure(command=do_save)
        entry.bind("<Return>", lambda e: do_save())
        win.bind("<Return>", lambda e: do_save())
        win.bind("<KP_Enter>", lambda e: do_save())
        win.bind("<Escape>", lambda e: win.destroy())

        # Centrar sobre la ventana principal.
        win.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() - win.winfo_width()) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() - win.winfo_height()) // 3
        win.geometry(f"+{x}+{y}")
        win.grab_set()

    def _on_fob_saved(self, win, sku: str):
        win.destroy()
        self._flash_status(f"FOB guardado para {sku} ✓")
        # Refrescar el detalle por si la fila visible es de este SKU.
        self._on_select()

    def _on_alt_click_mult(self, event):
        row = self.tree.identify_row(event.y)
        if not row or self.tree.parent(row) == "":
            return "break"
        info = self.leaf_to_item.get(row)
        if not info:
            return "break"
        sku = info.get("sku") or ""
        if not sku:
            self._flash_status("Esta venta no tiene SKU — no se puede cargar multiplicador")
            return "break"
        if local_store.get_fob(sku) is None:
            self._flash_status(
                "Cargá primero el precio FOB (Ctrl+Click)"
            )
            return "break"
        self._open_mult_modal(sku, info.get("title") or "")
        return "break"

    def _open_mult_modal(self, sku: str, title: str):
        win = tk.Toplevel(self.root)
        win.title("Editar multiplicador")
        win.transient(self.root)
        win.resizable(False, False)

        frame = ttk.Frame(win, padding=15)
        frame.pack(fill="both", expand=True)

        ttk.Label(
            frame,
            text=title,
            wraplength=420,
            font=("TkDefaultFont", 11, "bold"),
        ).pack(anchor="w", pady=(0, 4))
        ttk.Label(frame, text=f"SKU: {sku}", foreground="#666").pack(
            anchor="w", pady=(0, 12)
        )

        ttk.Label(frame, text="Multiplicador (unidades por publicación):").pack(
            anchor="w"
        )
        actual = local_store.get_multiplicador(sku)
        mult_var = tk.StringVar(value=str(actual) if actual else "")
        entry = ttk.Entry(frame, textvariable=mult_var, width=12)
        entry.pack(anchor="w", pady=(2, 12))
        entry.focus_set()
        entry.select_range(0, "end")

        ttk.Label(
            frame,
            text="Entero ≥ 1. Para una publi de pack/12, poner 12.",
            foreground="#888",
            font=("TkDefaultFont", 9, "italic"),
        ).pack(anchor="w", pady=(0, 8))

        btns = ttk.Frame(frame)
        btns.pack(fill="x")

        save_btn = ttk.Button(btns, text="Guardar")
        cancel_btn = ttk.Button(btns, text="Cancelar", command=win.destroy)
        save_btn.pack(side="right")
        cancel_btn.pack(side="right", padx=(0, 6))

        def do_save():
            raw = mult_var.get().strip()
            try:
                new_mult = int(raw)
            except ValueError:
                self._flash_status("Valor inválido — debe ser un entero ≥ 1")
                return
            if new_mult < 1:
                self._flash_status("El multiplicador debe ser ≥ 1")
                return
            try:
                local_store.set_multiplicador(sku, new_mult)
            except Exception as e:
                messagebox.showerror(
                    "Error", f"No se pudo guardar el multiplicador:\n{e}", parent=win
                )
                return
            self._on_mult_saved(win, sku)

        save_btn.configure(command=do_save)
        entry.bind("<Return>", lambda e: do_save())
        win.bind("<Return>", lambda e: do_save())
        win.bind("<KP_Enter>", lambda e: do_save())
        win.bind("<Escape>", lambda e: win.destroy())

        win.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() - win.winfo_width()) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() - win.winfo_height()) // 3
        win.geometry(f"+{x}+{y}")
        win.grab_set()

    def _on_mult_saved(self, win, sku: str):
        win.destroy()
        self._flash_status(f"Multiplicador guardado para {sku} ✓")
        self._on_select()

    def _open_markup_modal(self, sku: str, title: str):
        """Modal para editar el markup de Andrés (×N) específico de un SKU.
        Se persiste en data.json. Si se borra (vacío) o se pone el default,
        vuelve a usar GANANCIA_HERMANO_MULT (= 1.3) para mantener el JSON limpio."""
        if not sku:
            self._flash_status("Seleccioná una venta con SKU primero")
            return
        win = tk.Toplevel(self.root)
        win.title("Markup de Andrés")
        win.transient(self.root)
        win.resizable(False, False)

        frame = ttk.Frame(win, padding=15)
        frame.pack(fill="both", expand=True)

        if title:
            ttk.Label(
                frame,
                text=title,
                wraplength=420,
                font=("TkDefaultFont", 11, "bold"),
            ).pack(anchor="w", pady=(0, 4))
        ttk.Label(frame, text=f"SKU: {sku}", foreground="#666").pack(
            anchor="w", pady=(0, 12)
        )

        ttk.Label(frame, text="Markup de Andrés:").pack(anchor="w")
        actual = local_store.get_markup(sku)
        if actual is None:
            actual = GANANCIA_HERMANO_MULT
        markup_var = tk.StringVar(value=f"{actual:g}")
        entry = ttk.Entry(frame, textvariable=markup_var, width=12)
        entry.pack(anchor="w", pady=(2, 12))
        entry.focus_set()
        entry.select_range(0, "end")

        ttk.Label(
            frame,
            text=(
                f"Default {GANANCIA_HERMANO_MULT:g}. Subilo si Andrés te cobra "
                f"más por este producto. Vacío o {GANANCIA_HERMANO_MULT:g} = default."
            ),
            foreground="#888",
            font=("TkDefaultFont", 9, "italic"),
            wraplength=380,
        ).pack(anchor="w", pady=(0, 8))

        btns = ttk.Frame(frame)
        btns.pack(fill="x")

        def do_save():
            # Punto como separador decimal. Acepta coma como tolerancia.
            raw = markup_var.get().strip().replace(",", ".")
            if raw == "":
                new_markup = None  # borra el override
            else:
                try:
                    new_markup = float(raw)
                except ValueError:
                    self._flash_status("Valor inválido — debe ser un número ≥ 1")
                    return
                if new_markup < 1:
                    self._flash_status("El markup debe ser ≥ 1")
                    return
                # Si volvió al default, borramos el override para mantener JSON limpio.
                if abs(new_markup - GANANCIA_HERMANO_MULT) < 1e-9:
                    new_markup = None
            try:
                local_store.set_markup(sku, new_markup)
            except Exception as e:
                messagebox.showerror(
                    "Error", f"No se pudo guardar el markup:\n{e}", parent=win
                )
                return
            win.destroy()
            shown = (
                f"{GANANCIA_HERMANO_MULT:g} (default)"
                if new_markup is None
                else f"{new_markup:g}"
            )
            self._flash_status(f"Markup ×{shown} guardado para {sku} ✓")
            self._on_select()
            self._update_totales_inline()

        ttk.Button(btns, text="Guardar", command=do_save).pack(side="right")
        ttk.Button(btns, text="Cancelar", command=win.destroy).pack(
            side="right", padx=(0, 6)
        )
        entry.bind("<Return>", lambda _e: do_save())
        win.bind("<Return>", lambda _e: do_save())
        win.bind("<KP_Enter>", lambda _e: do_save())
        win.bind("<Escape>", lambda _e: win.destroy())

        win.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() - win.winfo_width()) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() - win.winfo_height()) // 3
        win.geometry(f"+{x}+{y}")
        win.grab_set()

    def _open_totales_modal(self):
        totales = self._calcular_totales_seleccionados()
        selected_total = local_store.count_checked()

        win = tk.Toplevel(self.root)
        win.title("Totales de seleccionadas")
        win.transient(self.root)

        outer = ttk.Frame(win, padding=15)
        outer.pack(fill="both", expand=True)

        # Header
        header_text = f"{totales['count_total']} ventas seleccionadas cargadas"
        if selected_total > totales["count_total"]:
            header_text += (
                f"  ·  {selected_total - totales['count_total']} más en historial "
                f"no cargado"
            )
        ttk.Label(
            outer, text=header_text, font=("TkDefaultFont", 11, "bold")
        ).pack(anchor="w", pady=(0, 10))

        if totales["count_total"] == 0:
            ttk.Label(
                outer,
                text="No hay ventas seleccionadas en las cargadas.",
                foreground="#888",
            ).pack(anchor="w")
            ttk.Button(outer, text="Cerrar", command=win.destroy).pack(
                anchor="e", pady=(15, 0)
            )
            win.bind("<Escape>", lambda e: win.destroy())
            return

        # Bloque de números
        numbers = ttk.Frame(outer)
        numbers.pack(fill="x", pady=(0, 10))

        def add_row(parent, label, value, color, bold=False, big=False):
            row = ttk.Frame(parent)
            row.pack(fill="x", anchor="w", pady=2)
            size = 12 if big else 10
            font = ("TkDefaultFont", size, "bold") if bold else ("TkDefaultFont", size)
            ttk.Label(row, text=label, font=font).pack(side="left")
            tk.Label(
                row, text=value, foreground=color, font=font
            ).pack(side="right")

        add_row(numbers, "Bruto MP", format_price(totales["bruto"]), "#1a3a5c")
        add_row(numbers, "Neto MP", format_price(totales["neto_all"]), "#1e7a1e")

        if totales["count_calc"] > 0:
            ttk.Separator(numbers, orient="horizontal").pack(fill="x", pady=6)

            # Centramos las dos filas con grid: columnas 0 y 2 absorben
            # el espacio sobrante, la fila va en la columna 1 (centrada).
            center_box = ttk.Frame(numbers)
            center_box.pack(fill="x", pady=(2, 0))
            center_box.grid_columnconfigure(0, weight=1)
            center_box.grid_columnconfigure(2, weight=1)

            costo_text = f"Total para pagar a Andrés - {format_price(totales['costo'])}"
            ganancia_text = (
                f"Ganancia de Pablo total {format_price(totales['ganancia'])}"
            )

            def make_clickable_row(parent, grid_row, label_text, value_text, color, copy_text):
                row_frame = ttk.Frame(parent)
                row_frame.grid(row=grid_row, column=1, pady=2)
                lbl_left = tk.Label(
                    row_frame,
                    text=label_text,
                    foreground=color,
                    font=("TkDefaultFont", 12, "bold"),
                    cursor="hand2",
                )
                lbl_left.pack(side="left", padx=(0, 6))
                lbl_right = tk.Label(
                    row_frame,
                    text=value_text,
                    foreground=color,
                    font=("TkDefaultFont", 12, "bold"),
                    cursor="hand2",
                )
                lbl_right.pack(side="left")

                def on_click(_e=None):
                    self._set_clipboard(copy_text)
                    self._flash_status(f"Copiado: {copy_text}")

                for w in (row_frame, lbl_left, lbl_right):
                    w.bind("<Button-1>", on_click)

            make_clickable_row(
                center_box, 0,
                "Total para pagar a Andrés →",
                f"- {format_price(totales['costo'])}",
                "#c0392b",
                costo_text,
            )
            make_clickable_row(
                center_box, 1,
                "Ganancia de Pablo total →",
                format_price(totales["ganancia"]),
                "#1e7a1e",
                ganancia_text,
            )

        # Listado de problemas
        sin_costo = totales["sin_costo"]
        if sin_costo:
            ttk.Separator(outer, orient="horizontal").pack(fill="x", pady=(10, 8))
            ttk.Label(
                outer,
                text=f"⚠ {len(sin_costo)} sin costo cargado:",
                foreground="#d35400",
                font=("TkDefaultFont", 10, "bold"),
            ).pack(anchor="w", pady=(0, 4))

            # Listbox con scroll por si hay muchos
            list_frame = ttk.Frame(outer)
            list_frame.pack(fill="both", expand=True)
            lst = tk.Listbox(
                list_frame,
                height=min(8, len(sin_costo)),
                width=70,
                font=("TkDefaultFont", 9),
                activestyle="none",
            )
            lst_vsb = ttk.Scrollbar(
                list_frame, orient="vertical", command=lst.yview
            )
            lst.configure(yscrollcommand=lst_vsb.set)
            lst.pack(side="left", fill="both", expand=True)
            lst_vsb.pack(side="right", fill="y")

            # Agrupar por SKU para no repetir si hay varias ventas del mismo
            seen: dict[str, tuple[str, str]] = {}
            for sku, title, motivo in sin_costo:
                key = f"{sku}|{motivo}"
                if key not in seen:
                    seen[key] = (title, motivo)
            for key, (title, motivo) in seen.items():
                sku = key.split("|", 1)[0]
                sku_part = sku if sku else "(sin SKU)"
                title_part = title[:50] + ("…" if len(title) > 50 else "")
                lst.insert("end", f"  [{motivo}]  {sku_part}  —  {title_part}")

        ttk.Button(outer, text="Cerrar", command=win.destroy).pack(
            anchor="e", pady=(15, 0)
        )
        win.bind("<Escape>", lambda e: win.destroy())

        win.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() - win.winfo_width()) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() - win.winfo_height()) // 3
        win.geometry(f"+{x}+{y}")

    def _cargar_dolar_async(self):
        def worker():
            dolar.cargar()
            self.root.after(0, self._on_dolar_cargado)
        threading.Thread(target=worker, daemon=True).start()

    # ────────────── Informe para WhatsApp ──────────────

    def _copy_informe_wasap(self):
        """Arma un informe explicativo de la venta seleccionada con todos los
        cálculos y lo copia al portapapeles. Formato WhatsApp (negritas con *,
        emojis, párrafos cortos). Pensado para mandárselo a Andrés/Pablo y que
        cualquiera lo entienda sin abrir la app."""
        sel = self.tree.selection()
        if not sel:
            self._flash_status("Seleccioná una venta primero")
            return
        leaf_id = sel[0]
        info = self.leaf_to_item.get(leaf_id)
        if not info:
            self._flash_status("Seleccioná una venta (no un día)")
            return
        order_id = self.row_to_order.get(leaf_id)

        title = info.get("title") or "(sin título)"
        sku = info.get("sku") or "(sin SKU)"
        try:
            quantity = int(info.get("quantity") or 1)
        except (TypeError, ValueError):
            quantity = 1
        bruto = float(info.get("total_amount") or 0)

        fob = local_store.get_fob(sku) if sku else None
        mult = local_store.get_multiplicador(sku) if sku else None
        markup = local_store.get_markup(sku) if sku else None
        if markup is None:
            markup = GANANCIA_HERMANO_MULT
        cot = dolar.get()
        neto = local_store.get_neto_efectivo(order_id) if order_id else None
        neto_raw = local_store.get_neto_manual(order_id) if order_id else None
        envio = local_store.get_shipping_manual(order_id) if order_id else None

        lineas: list[str] = []
        lineas.append(f"📦 *{title}*")
        lineas.append(f"🏷️ SKU: `{sku}`")
        if quantity > 1:
            lineas.append(f"🛒 Cantidad vendida: *{quantity} unidades*")
        else:
            lineas.append("🛒 Cantidad vendida: *1 unidad*")
        lineas.append(f"💵 Precio publicado: *{format_price(bruto)}*")
        lineas.append("")

        # ── Costo de importación ──
        if fob and mult and cot:
            nac_unit = fob * NACIONALIZACION_MULT
            pesos_unit = nac_unit * cot
            andres_unit = pesos_unit * markup
            pesos_pack = pesos_unit * mult
            andres_pack = andres_unit * mult
            pesos_total = pesos_pack * quantity
            andres_total = andres_pack * quantity

            lineas.append("🇨🇳 *COSTO DE IMPORTACIÓN (lo que pagó Andrés en China)*")
            lineas.append(f"• FOB unitario: USD {fob:,.2f}")
            lineas.append(
                f"• Nacionalizado (×{NACIONALIZACION_MULT}): USD {nac_unit:,.2f}"
            )
            lineas.append(
                f"• En pesos (cotización ${cot:,.0f}): {format_price(pesos_unit)} por unidad"
            )
            if mult > 1:
                lineas.append(
                    f"• Por pack de {mult}: {format_price(pesos_pack)}"
                )
            if quantity > 1:
                lineas.append(
                    f"• Total por las {quantity} unidades: *{format_price(pesos_total)}*"
                )
            lineas.append("")

            # ── Lo que cobra Andrés a Pablo ──
            markup_str = f"{markup:g}"
            lineas.append(
                f"💰 *LO QUE COBRA ANDRÉS A PABLO* (markup ×{markup_str} = {(markup-1)*100:.0f}%)"
            )
            lineas.append(
                f"• Por unidad: {format_price(andres_unit)}"
            )
            if mult > 1:
                lineas.append(
                    f"• Por pack de {mult}: {format_price(andres_pack)}"
                )
            if quantity > 1:
                lineas.append(
                    f"• Total a pagar a Andrés por esta venta: *{format_price(andres_total)}*"
                )
            lineas.append("")

            # ── Ganancia de Andrés ──
            gan_a_unit = andres_unit - pesos_unit
            gan_a_pack = gan_a_unit * mult
            gan_a_total = gan_a_pack * quantity
            margen_a_pct = (markup - 1) * 100

            lineas.append("👨 *GANANCIA DE ANDRÉS*")
            lineas.append(
                f"• Por unidad: {format_price(gan_a_unit)}"
            )
            if mult > 1:
                lineas.append(
                    f"• Por pack de {mult}: {format_price(gan_a_pack)}"
                )
            if quantity > 1:
                lineas.append(
                    f"• Total de esta venta: *{format_price(gan_a_total)}*"
                )
            lineas.append(
                f"• Margen sobre su costo: *{margen_a_pct:.1f}%*"
            )
            lineas.append("")

        # ── Cobro Mercado Pago ──
        lineas.append("💳 *COBRO EN MERCADO PAGO*")
        lineas.append(f"• Bruto (precio publicado): {format_price(bruto)}")
        if neto_raw is not None:
            lineas.append(
                f"• Neto MP (después de comisiones): {format_price(neto_raw)}"
            )
            if envio:
                lineas.append(
                    f"• Envío Flex (lo que pago afuera): -{format_price(envio)}"
                )
                lineas.append(
                    f"• Neto efectivo: *{format_price(neto)}*"
                )
        else:
            lineas.append("• ⚠️ Neto MP todavía no cargado")
        lineas.append("")

        # ── Ganancia total Pablo ──
        if (
            fob and mult and cot and neto is not None
        ):
            costo_unit_pack = pesos_unit * mult * markup  # what Pablo paga a Andrés (per pack)
            costo_total_pablo = costo_unit_pack * quantity
            ganancia_pablo = float(neto) - costo_total_pablo
            margen_pct = (
                (ganancia_pablo / bruto * 100) if bruto > 0 else 0
            )
            lineas.append("🎯 *RESULTADO FINAL PABLO*")
            lineas.append(
                f"• Le entró (neto efectivo): {format_price(neto)}"
            )
            lineas.append(
                f"• Le pagó a Andrés: -{format_price(costo_total_pablo)}"
            )
            signo = "✅" if ganancia_pablo >= 0 else "❌"
            lineas.append(
                f"• {signo} *Ganancia neta: {format_price(ganancia_pablo)}*"
            )
            lineas.append(
                f"• Margen sobre el bruto: *{margen_pct:.1f}%*"
            )

            # Niveles
            def nivel(pct):
                if pct < 10:
                    return "🔴 MUY BAJO"
                if pct < 20:
                    return "🟠 BAJO"
                if pct < 30:
                    return "🟢 BUENO"
                return "🟣 EXCELENTE"

            lineas.append(
                f"• Nivel: {nivel(margen_pct)}"
            )
            lineas.append("")
            lineas.append(
                "_Niveles: <10% MUY BAJO · 10-20% BAJO · 20-30% BUENO · >30% EXCELENTE_"
            )

        texto = "\n".join(lineas)
        self._set_clipboard(texto)
        self._flash_status("📋 Informe copiado al portapapeles ✓")

    def _copy_informe_lite(self):
        """Versión resumida del informe: solo título, SKU, cantidad, ganancia
        Andrés y ganancia Pablo. Pensado para mandar rápido por WhatsApp sin
        toda la explicación de los cálculos."""
        sel = self.tree.selection()
        if not sel:
            self._flash_status("Seleccioná una venta primero")
            return
        leaf_id = sel[0]
        info = self.leaf_to_item.get(leaf_id)
        if not info:
            self._flash_status("Seleccioná una venta (no un día)")
            return
        order_id = self.row_to_order.get(leaf_id)

        title = info.get("title") or "(sin título)"
        sku = info.get("sku") or "(sin SKU)"
        try:
            quantity = int(info.get("quantity") or 1)
        except (TypeError, ValueError):
            quantity = 1

        fob = local_store.get_fob(sku) if sku else None
        mult = local_store.get_multiplicador(sku) if sku else None
        markup = local_store.get_markup(sku) if sku else None
        if markup is None:
            markup = GANANCIA_HERMANO_MULT
        cot = dolar.get()
        neto = local_store.get_neto_efectivo(order_id) if order_id else None

        lineas: list[str] = []
        lineas.append(f"📦 *{title}*")
        lineas.append(f"🏷️ SKU: `{sku}`")
        unidad_label = "unidad" if quantity == 1 else "unidades"
        lineas.append(f"🛒 Cantidad vendida: *{quantity} {unidad_label}*")
        lineas.append("")

        if fob and mult and cot:
            pesos_unit = fob * NACIONALIZACION_MULT * cot
            andres_total = pesos_unit * markup * mult * quantity
            pesos_total = pesos_unit * mult * quantity
            gan_andres = andres_total - pesos_total
            lineas.append(f"👨 *Ganancia Andrés:* {format_price(gan_andres)}")

            if neto is not None:
                gan_pablo = float(neto) - andres_total
                signo = "✅" if gan_pablo >= 0 else "❌"
                lineas.append(f"{signo} *Ganancia Pablo:* {format_price(gan_pablo)}")
            else:
                lineas.append("⚠️ *Ganancia Pablo:* falta cargar neto MP")
        else:
            lineas.append("⚠️ Faltan datos (FOB / multiplicador / dólar) para calcular ganancias")

        texto = "\n".join(lineas)
        self._set_clipboard(texto)
        self._flash_status("📄 Informe lite copiado al portapapeles ✓")

    # ────────────── Frase del día (modal) ──────────────
    # Tk en Linux no rendea emojis en color con la fuente default. Symbola
    # tiene Latin + emojis (mono) en una sola fuente, así que renderizamos
    # la frase con PIL y la mostramos como PhotoImage. Sin esto los emojis
    # aparecen como cuadraditos vacíos. Ver skill `iconos-pil-tkinter`.
    _SYMBOLA_PATH = "/usr/share/fonts/gdouros-symbola/Symbola.ttf"
    _MONO_PATH = "/usr/share/fonts/liberation-mono-fonts/LiberationMono-Bold.ttf"

    def _open_frase_modal(self):
        win = tk.Toplevel(self.root)
        win.title("Frase del día")
        win.transient(self.root)
        win.resizable(False, False)

        outer = ttk.Frame(win, padding=20)
        outer.pack(fill="both", expand=True)

        ttk.Label(
            outer,
            text="✨ Frase del día de Claude ✨",
            font=("TkDefaultFont", 14, "bold"),
            foreground="#7d3c98",
        ).pack(anchor="w", pady=(0, 12))

        body = ttk.Frame(outer)
        body.pack(fill="both", expand=True, pady=(0, 12))

        # Label que vamos a actualizar (texto loading → imagen frase).
        frase_lbl = ttk.Label(
            body,
            text="Cargando frase…",
            foreground="#888",
            font=("TkDefaultFont", 11, "italic"),
        )
        frase_lbl.pack(anchor="w")

        btn_bar = ttk.Frame(outer)
        btn_bar.pack(fill="x")

        send_btn = ttk.Button(btn_bar, text="📨 Mandar a Pablo")
        send_btn.pack(side="left")
        send_btn.configure(state="disabled")

        refresh_btn = ttk.Button(btn_bar, text="🔄 Otra")
        refresh_btn.pack(side="left", padx=(8, 0))
        refresh_btn.configure(state="disabled")

        ttk.Button(btn_bar, text="Cerrar", command=win.destroy).pack(side="right")

        status_lbl = ttk.Label(
            outer, text="", foreground="#888",
            font=("TkDefaultFont", 9, "italic"),
        )
        status_lbl.pack(anchor="w", pady=(8, 0))

        # Mantenemos refs a la imagen para que el GC no se la coma.
        state = {"img": None, "texto": None}

        def cargar():
            send_btn.configure(state="disabled")
            refresh_btn.configure(state="disabled")
            frase_lbl.configure(image="", text="Cargando frase…")
            status_lbl.configure(text="")

            def worker():
                texto = frase.cargar()
                self.root.after(0, lambda: on_loaded(texto))
            threading.Thread(target=worker, daemon=True).start()

        def on_loaded(texto):
            if not texto:
                frase_lbl.configure(image="", text="(no se pudo cargar)")
                refresh_btn.configure(state="normal")
                return
            state["texto"] = texto
            try:
                photo = self._render_frase_image(texto, max_width=520)
            except Exception as e:
                frase_lbl.configure(image="", text=f"(error rindiendo: {e})")
                refresh_btn.configure(state="normal")
                return
            state["img"] = photo
            frase_lbl.configure(image=photo, text="")
            send_btn.configure(state="normal")
            refresh_btn.configure(state="normal")

        def enviar():
            texto = state["texto"]
            if not texto:
                return
            send_btn.configure(state="disabled")
            status_lbl.configure(text="Enviando…", foreground="#888")

            def worker():
                ok, detalle = whatsapp_send.enviar(texto)
                self.root.after(0, lambda: on_sent(ok, detalle))
            threading.Thread(target=worker, daemon=True).start()

        def on_sent(ok, detalle):
            if ok:
                status_lbl.configure(text="✓ Enviado a Pablo", foreground="#1e7a1e")
            else:
                status_lbl.configure(
                    text=f"✕ {detalle[:80]}", foreground="#c0392b"
                )
            send_btn.configure(state="normal")

        send_btn.configure(command=enviar)
        refresh_btn.configure(command=cargar)
        win.bind("<Escape>", lambda e: win.destroy())

        # Centrar respecto a la ventana principal.
        win.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() - win.winfo_width()) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() - win.winfo_height()) // 3
        win.geometry(f"+{x}+{y}")

        cargar()

    def _render_frase_image(self, texto: str, max_width: int = 520):
        """Rinde la frase con look CRT verde fósforo: mono bold para texto,
        Symbola para emojis (que mono no tiene), fondo negro."""
        from PIL import Image, ImageDraw, ImageFont, ImageTk

        font_size = 20
        color = (51, 255, 51)   # verde fósforo CRT (#33ff33)
        bg = (0, 0, 0)          # negro
        padding = 18

        mono = ImageFont.truetype(self._MONO_PATH, font_size)
        emoji = ImageFont.truetype(self._SYMBOLA_PATH, font_size)

        def font_for(ch: str):
            # Heurística: ASCII + Latin-1 + signos básicos van en mono.
            # Todo lo "raro" (emojis, símbolos, dingbats) cae a Symbola.
            cp = ord(ch)
            if cp < 0x2000:
                return mono
            return emoji

        def measure(s: str) -> int:
            # Medimos sumando el ancho de cada char con su fuente.
            w = 0
            for ch in s:
                bbox = font_for(ch).getbbox(ch)
                w += bbox[2] - bbox[0] if bbox else 0
            return w

        # Wrap manual por palabras.
        max_text_width = max_width - 2 * padding
        words = texto.split()
        lines: list[str] = []
        current = ""
        for w in words:
            candidate = (current + " " + w).strip()
            if measure(candidate) <= max_text_width or not current:
                current = candidate
            else:
                lines.append(current)
                current = w
        if current:
            lines.append(current)

        ascent, descent = mono.getmetrics()
        line_h = ascent + descent + 6
        height = padding * 2 + line_h * len(lines)
        width = max_width

        img = Image.new("RGB", (width, height), bg)
        draw = ImageDraw.Draw(img)
        y = padding
        for line in lines:
            x = padding
            for ch in line:
                f = font_for(ch)
                draw.text((x, y), ch, font=f, fill=color)
                bbox = f.getbbox(ch)
                x += (bbox[2] - bbox[0]) if bbox else 0
            y += line_h

        return ImageTk.PhotoImage(img)
    # ──────────────────────────────────────────────────


    def _on_dolar_cargado(self):
        cot = dolar.get()
        if cot is None:
            self.dolar_var.set("USD ✕")
        else:
            self.dolar_var.set(f"USD ${cot:,.0f}")
        # Refrescar el detalle por si ya hay una fila seleccionada.
        self._on_select()

    def _all_leaves(self) -> list:
        leaves = []
        for parent in self.tree.get_children(""):
            leaves.extend(self.tree.get_children(parent))
        return leaves

    def _set_loading(self, loading: bool):
        self.loading = loading
        state = "disabled" if loading else "normal"
        self.btn_refresh.configure(state=state)
        if loading:
            self.btn_more.configure(state="disabled")
            self.status_var.set("Cargando...")
        else:
            self._update_status()
            shown = len(self._all_leaves())
            if shown < self.total:
                self.btn_more.configure(state="normal")
            else:
                self.btn_more.configure(state="disabled")

    # ──────────────────── Filtro del Treeview ────────────────────

    def _setup_filter_placeholder(
        self, entry: ttk.Entry, var: tk.StringVar, placeholder: str
    ) -> None:
        """Placeholder gris que desaparece al hacer focus."""
        def _on_focus_in(_e):
            if entry.get() == placeholder:
                entry.delete(0, "end")
                entry.configure(foreground="")

        def _on_focus_out(_e):
            if not entry.get():
                entry.insert(0, placeholder)
                entry.configure(foreground="#888")

        entry.insert(0, placeholder)
        entry.configure(foreground="#888")
        entry.bind("<FocusIn>", _on_focus_in)
        entry.bind("<FocusOut>", _on_focus_out)

    def _get_filter_text(self, var: tk.StringVar, placeholder: str) -> str:
        """Texto del filtro ignorando el placeholder."""
        val = var.get()
        return "" if val == placeholder else val

    def _refresh_leaf_meta(self, leaf_id: str):
        """Recalcula el row_text del filtro a partir del info actual del leaf."""
        meta = self._leaves_meta.get(leaf_id)
        if not meta:
            return
        info = self.leaf_to_item.get(leaf_id) or {}
        sku = info.get("sku") or ""
        title = info.get("title") or ""
        meta["row_text"] = _normalize(f"{sku} {title}")

    def _limpiar_filtros(self):
        # Limpiar SIN disparar el placeholder primero — seteamos en vacío y
        # forzamos el placeholder solo si el entry no tiene focus.
        for entry, var, placeholder in (
            (self._buscar_entry, self._buscar_var, self._BUSCAR_PLACEHOLDER),
            (self._excluir_entry, self._excluir_var, self._EXCLUIR_PLACEHOLDER),
        ):
            var.set("")
            if self.root.focus_get() is not entry:
                entry.insert(0, placeholder)
                entry.configure(foreground="#888")
        # También apagar el toggle de "solo con nota" si está prendido.
        if self._solo_con_nota_var.get():
            self._solo_con_nota_var.set(False)

    def _refresh_tree_filter(self):
        """Aplica los filtros actuales sobre las leaves cargadas (detach/move)."""
        # Guard: este callback puede dispararse durante _build_ui (vía
        # trace_add de los filter vars cuando _setup_filter_placeholder
        # set()ea el placeholder) antes de que self.tree exista. Si no
        # hay tree todavía, simplemente no hacemos nada.
        if not hasattr(self, "tree"):
            return
        buscar_raw = self._get_filter_text(self._buscar_var, self._BUSCAR_PLACEHOLDER)
        excluir_raw = self._get_filter_text(
            self._excluir_var, self._EXCLUIR_PLACEHOLDER
        )
        buscar_words = [_normalize(w) for w in buscar_raw.split() if w.strip()]
        excluir_words = [_normalize(w) for w in excluir_raw.split() if w.strip()]
        solo_con_nota = self._solo_con_nota_var.get()
        filter_active = bool(buscar_words or excluir_words or solo_con_nota)

        # Decidir qué leaves quedan visibles, agrupados por día y en orden original.
        visible_per_day: dict[str, list] = {}
        for leaf_id, meta in sorted(
            self._leaves_meta.items(), key=lambda kv: kv[1]["order"]
        ):
            row_text = meta["row_text"]
            if buscar_words and not all(w in row_text for w in buscar_words):
                continue
            if excluir_words and all(w in row_text for w in excluir_words):
                continue
            if solo_con_nota:
                order_id = self.row_to_order.get(leaf_id)
                if not order_id or not local_store.has_nota(order_id):
                    continue
            visible_per_day.setdefault(meta["day_key"], []).append(leaf_id)

        # Detach todas las leaves de todas las días (sin tocar los días en sí).
        for day_parent_id in self.day_nodes.values():
            for leaf_id in list(self.tree.get_children(day_parent_id)):
                self.tree.detach(leaf_id)

        # Reattach las visibles, en orden original dentro de cada día.
        for day_key, leaf_ids in visible_per_day.items():
            parent_id = self.day_nodes.get(day_key)
            if not parent_id:
                continue
            for leaf_id in leaf_ids:
                self.tree.move(leaf_id, parent_id, "end")

        # Detach días vacíos. Reordenar los visibles por fecha desc.
        sorted_days = sorted(
            self.day_nodes.keys(),
            key=lambda d: datetime.strptime(d, "%d/%m/%Y") if d else datetime.min,
            reverse=True,
        )
        for day_key in sorted_days:
            parent_id = self.day_nodes[day_key]
            if self.tree.get_children(parent_id):
                self.tree.move(parent_id, "", "end")
            else:
                self.tree.detach(parent_id)

        # Refrescar checks y header del check global (puede cambiar el "todos marcados").
        for parent in self.tree.get_children(""):
            self._update_day_check(parent)
        self._update_header_check()

        # Status: cuando hay filtro, mostrar X de Y mostradas.
        self._filter_active = filter_active
        self._update_status()

    def _calcular_totales_seleccionados(self) -> dict:
        """Itera ventas cargadas y checkeadas, devuelve totales.

        Solo incluye en el costo/ganancia las ventas con FOB + multiplicador
        + cotización del dólar + neto MP manual disponibles. Las que no, van
        a `sin_costo` con un motivo para mostrar en el modal de detalle.
        """
        bruto = 0.0
        neto_all = 0.0
        neto_calc = 0.0
        costo = 0.0
        sin_costo: list[tuple[str, str, str]] = []
        count_total = 0
        count_calc = 0
        cot = dolar.get()

        # Iterar TODAS las leaves cargadas (incluso ocultas por el filtro),
        # así los totales son independientes del filtro visual.
        for leaf_id, order_id in self.row_to_order.items():
            if not order_id or not local_store.is_checked(order_id):
                continue
            info = self.leaf_to_item.get(leaf_id) or {}
            count_total += 1
            bruto += float(info.get("total_amount") or 0)
            # Neto efectivo: lo que queda después de restar Flex (si aplica).
            neto_manual = local_store.get_neto_efectivo(order_id)
            if neto_manual is not None:
                neto_all += neto_manual

            sku = info.get("sku") or ""
            title = info.get("title") or ""
            try:
                quantity = int(info.get("quantity") or 1)
            except (TypeError, ValueError):
                quantity = 1

            if neto_manual is None:
                sin_costo.append((sku, title, "sin neto MP"))
                continue
            if not sku:
                sin_costo.append(("", title, "sin SKU"))
                continue
            fob = local_store.get_fob(sku)
            if not fob or fob <= 0:
                sin_costo.append((sku, title, "sin FOB"))
                continue
            mult = local_store.get_multiplicador(sku)
            if mult is None:
                sin_costo.append((sku, title, "sin multiplicador"))
                continue
            if cot is None:
                sin_costo.append((sku, title, "sin cotización dólar"))
                continue

            markup = local_store.get_markup(sku) or GANANCIA_HERMANO_MULT
            costo_unit = (
                fob * mult * NACIONALIZACION_MULT * cot * markup
            )
            costo += costo_unit * quantity
            neto_calc += neto_manual
            count_calc += 1

        return {
            "count_total": count_total,
            "count_calc": count_calc,
            "bruto": bruto,
            "neto_all": neto_all,
            "neto_calc": neto_calc,
            "costo": costo,
            "ganancia": neto_calc - costo,
            "sin_costo": sin_costo,
        }

    def _update_status(self):
        loaded = len(self._leaves_meta)
        days = len(self.tree.get_children(""))
        if self._filter_active:
            visible = len(self._all_leaves())
            self.status_var.set(
                f"{visible} de {loaded} cargadas (filtro)  •  "
                f"{local_store.count_checked()} seleccionadas"
            )
        else:
            self.status_var.set(
                f"{loaded} de {self.total} ventas en {days} días  •  "
                f"{local_store.count_checked()} seleccionadas"
            )
        self._update_totales_inline()

    def _update_totales_inline(self):
        """Refresca el mini totalizador de la barra inferior."""
        totales = self._calcular_totales_seleccionados()
        if totales["count_total"] == 0:
            self.totales_costo_var.set("")
            self.totales_ganancia_var.set("")
            return
        self.totales_costo_var.set(f"Costo: {format_price(totales['costo'])}")
        self.totales_ganancia_var.set(
            f"Ganancia: {format_price(totales['ganancia'])}"
        )
        self.totales_ganancia_lbl.configure(
            foreground="#1e7a1e" if totales["ganancia"] >= 0 else "#c0392b"
        )

    def _set_leaf_check(self, leaf_id: str, checked: bool):
        order_id = self.row_to_order.get(leaf_id)
        if not order_id:
            return
        local_store.set_check(order_id, checked)
        values = list(self.tree.item(leaf_id, "values"))
        values[0] = CHECKED if checked else UNCHECKED
        self.tree.item(leaf_id, values=values, tags=self._row_tags(leaf_id, order_id))

    def _toggle(self, leaf_id: str):
        order_id = self.row_to_order.get(leaf_id)
        if order_id is None:
            return
        self._set_leaf_check(leaf_id, not local_store.is_checked(order_id))
        parent = self.tree.parent(leaf_id)
        if parent:
            self._update_day_check(parent)
        self._update_header_check()
        self._update_status()

    def _toggle_day(self, parent_id: str):
        leaves = self.tree.get_children(parent_id)
        if not leaves:
            return
        all_checked = all(
            local_store.is_checked(self.row_to_order.get(l) or "") for l in leaves
        )
        for l in leaves:
            self._set_leaf_check(l, not all_checked)
        self._update_day_check(parent_id)
        self._update_header_check()
        self._update_status()

    def _toggle_all(self):
        leaves = self._all_leaves()
        if not leaves:
            return
        all_checked = all(
            local_store.is_checked(self.row_to_order.get(l) or "") for l in leaves
        )
        for l in leaves:
            self._set_leaf_check(l, not all_checked)
        for parent in self.tree.get_children(""):
            self._update_day_check(parent)
        self._update_header_check()
        self._update_status()

    def _update_day_check(self, parent_id: str):
        leaves = self.tree.get_children(parent_id)
        if not leaves:
            return
        all_checked = all(
            local_store.is_checked(self.row_to_order.get(l) or "") for l in leaves
        )
        values = list(self.tree.item(parent_id, "values"))
        values[0] = CHECKED if all_checked else UNCHECKED
        self.tree.item(parent_id, values=values)

    def _update_header_check(self):
        leaves = self._all_leaves()
        if leaves and all(
            local_store.is_checked(self.row_to_order.get(l) or "") for l in leaves
        ):
            self.tree.heading("check", text=CHECKED)
        else:
            self.tree.heading("check", text=UNCHECKED)

    def _row_tags(self, row_id: str, order_id: str) -> tuple:
        base = self.row_base.get(row_id, "even")
        tags: list[str] = ["selected"] if local_store.is_checked(order_id) else [base]
        if local_store.has_nota(order_id):
            tags.append("with_note")
        return tuple(tags)

    def _on_right_click(self, event):
        row = self.tree.identify_row(event.y)
        self._right_clicked_row = row
        # Habilitar/deshabilitar items que requieren una hoja.
        is_leaf = bool(row) and self.tree.parent(row) != ""
        leaf_state = "normal" if is_leaf else "disabled"
        self.context_menu.entryconfig("Copiar título", state=leaf_state)
        self.context_menu.entryconfig("Copiar SKU", state=leaf_state)
        self.context_menu.entryconfig("Copiar ID publicación", state=leaf_state)
        self.context_menu.tk_popup(event.x_root, event.y_root)
        self.context_menu.focus_set()

    def _on_double_click(self, event):
        row = self.tree.identify_row(event.y)
        if not row or self.tree.parent(row) == "":
            return
        info = self.leaf_to_item.get(row)
        if not info or not info.get("item_id"):
            return
        self._open_sku_modal(row, info)

    def _open_sku_modal(self, leaf_id: str, info: dict):
        win = tk.Toplevel(self.root)
        win.title("Editar SKU")
        win.transient(self.root)
        win.resizable(False, False)

        frame = ttk.Frame(win, padding=15)
        frame.pack(fill="both", expand=True)

        ttk.Label(
            frame,
            text=info.get("title", ""),
            wraplength=420,
            font=("TkDefaultFont", 11, "bold"),
        ).pack(anchor="w", pady=(0, 4))

        meta = info.get("item_id", "")
        if info.get("variation_id"):
            meta += f"  ·  variación {info['variation_id']}"
        ttk.Label(frame, text=meta, foreground="#666").pack(anchor="w", pady=(0, 12))

        ttk.Label(frame, text="SKU:").pack(anchor="w")
        sku_var = tk.StringVar(value=info.get("sku") or "")
        entry = ttk.Entry(frame, textvariable=sku_var, width=42)
        entry.pack(fill="x", pady=(2, 12))
        entry.focus_set()
        entry.select_range(0, "end")

        btns = ttk.Frame(frame)
        btns.pack(fill="x")

        save_btn = ttk.Button(btns, text="Guardar")
        cancel_btn = ttk.Button(btns, text="Cancelar", command=win.destroy)
        save_btn.pack(side="right")
        cancel_btn.pack(side="right", padx=(0, 6))

        def do_save():
            new_sku = sku_var.get().strip()
            if new_sku == (info.get("sku") or ""):
                win.destroy()
                return
            save_btn.configure(state="disabled", text="Guardando...")
            cancel_btn.configure(state="disabled")
            entry.configure(state="disabled")

            def worker():
                try:
                    self._put_item_sku(
                        info["item_id"], info.get("variation_id"), new_sku
                    )
                except HTTPError as e:
                    body = ""
                    try:
                        body = e.read().decode("utf-8", errors="replace")[:300]
                    except Exception:
                        pass
                    msg = f"HTTP {e.code}: {e.reason}\n{body}"
                    self.root.after(
                        0,
                        lambda m=msg: self._sku_save_failed(
                            win, save_btn, cancel_btn, entry, m
                        ),
                    )
                    return
                except Exception as e:
                    msg = str(e)
                    self.root.after(
                        0,
                        lambda m=msg: self._sku_save_failed(
                            win, save_btn, cancel_btn, entry, m
                        ),
                    )
                    return
                self.root.after(
                    0, lambda: self._on_sku_updated(leaf_id, info, new_sku, win)
                )

            threading.Thread(target=worker, daemon=True).start()

        save_btn.configure(command=do_save)
        entry.bind("<Return>", lambda e: do_save())
        win.bind("<Return>", lambda e: do_save())
        win.bind("<KP_Enter>", lambda e: do_save())
        win.bind("<Escape>", lambda e: win.destroy())

        # Centrar sobre la ventana principal.
        win.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - win.winfo_width()) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - win.winfo_height()) // 2
        win.geometry(f"+{x}+{y}")

        # En Wayland la ventana puede no estar "viewable" todavía cuando intentamos
        # hacer grab_set, lo que da TclError. Reintentamos hasta que esté visible.
        def _set_grab():
            try:
                win.grab_set()
            except tk.TclError:
                win.after(50, _set_grab)
        win.after(10, _set_grab)

    def _sku_save_failed(self, win, save_btn, cancel_btn, entry, msg: str):
        save_btn.configure(state="normal", text="Guardar")
        cancel_btn.configure(state="normal")
        entry.configure(state="normal")
        messagebox.showerror("Error", f"No se pudo actualizar SKU:\n{msg}", parent=win)

    def _on_sku_updated(self, leaf_id: str, info: dict, new_sku: str, win):
        info["sku"] = new_sku
        if leaf_id in self.tree.get_children(self.tree.parent(leaf_id)):
            values = list(self.tree.item(leaf_id, "values"))
            # values = (check, time, sku, title, price)
            values[2] = new_sku
            self.tree.item(leaf_id, values=values)
        self._refresh_leaf_meta(leaf_id)
        win.destroy()
        self._flash_status(f"SKU actualizado: {new_sku}")

    def _put_item_sku(self, item_id: str, variation_id, new_sku: str):
        item_url = f"https://api.mercadolibre.com/items/{item_id}"

        # Decidir si el SKU vive en la variación o en el item.
        target_var = None
        vid = None
        if variation_id:
            try:
                vid = int(variation_id)
            except (TypeError, ValueError):
                vid = variation_id
            item_data = self._get_item(item_id)
            for v in item_data.get("variations") or []:
                if v.get("id") == vid:
                    target_var = v
                    break
            print(f"\n[SKU UPDATE] item={item_id} variation={vid}", flush=True)
            if target_var:
                print(f"  variation.seller_sku = {target_var.get('seller_sku')!r}", flush=True)
                print(f"  variation.attributes = {target_var.get('attributes')!r}", flush=True)
                print(f"  variation.attribute_combinations = {target_var.get('attribute_combinations')!r}", flush=True)
            print(f"  → setting to: {new_sku!r}", flush=True)

        # La variación tiene SKU propio sólo si su array attributes existe.
        # Si attributes es None, las variaciones se diferencian sólo por
        # combinations (color/talle) y el SKU vive a nivel item.
        use_variation_path = bool(target_var and target_var.get("attributes"))

        if use_variation_path:
            attrs = list(target_var.get("attributes") or [])
            updated = False
            for i, attr in enumerate(attrs):
                if (attr.get("id") or "").upper() == "SELLER_SKU":
                    attrs[i] = {"id": "SELLER_SKU", "value_name": new_sku}
                    updated = True
                    break
            if not updated:
                attrs.append({"id": "SELLER_SKU", "value_name": new_sku})

            try:
                result = self._do_put(
                    item_url,
                    {"variations": [{"id": vid, "attributes": attrs}]},
                )
            except HTTPError as e:
                body_text = self._read_err_body(e)
                if e.code == 400 and "item.pictures.max" in body_text:
                    msg = (
                        "El item tiene más de 12 fotos y la categoría ya no "
                        "permite ese límite. Reducí las fotos desde el panel "
                        "de Mercado Libre y volvé a intentar.\n\n" + body_text
                    )
                else:
                    msg = body_text
                raise HTTPError(
                    e.url, e.code, e.reason, e.headers,
                    io.BytesIO(msg.encode("utf-8")),
                )
        elif target_var is not None:
            # Item con variaciones pero la variación no tiene attributes propios.
            # No podemos usar item.attributes[SELLER_SKU] porque ML rechaza con
            # "item.attributes.invalid: Same attributes are used in item and
            # variations". El único camino que funciona es seller_custom_field.
            print(f"  → item con variaciones sin SKU per-variación → seller_custom_field", flush=True)
            result = self._do_put(item_url, {"seller_custom_field": new_sku})
        else:
            # Item plano sin variaciones: probar attributes primero; si falla
            # por fotos, fallback a seller_custom_field.
            print(f"  → item plano, intentando attributes[SELLER_SKU]", flush=True)
            try:
                result = self._do_put(
                    item_url,
                    {"attributes": [{"id": "SELLER_SKU", "value_name": new_sku}]},
                )
            except HTTPError as e:
                body_text = self._read_err_body(e)
                if e.code == 400 and (
                    "item.pictures.max" in body_text
                    or "item.attributes.invalid" in body_text
                ):
                    print(f"  → fallback a seller_custom_field", flush=True)
                    result = self._do_put(
                        item_url, {"seller_custom_field": new_sku}
                    )
                else:
                    raise HTTPError(
                        e.url, e.code, e.reason, e.headers,
                        io.BytesIO(body_text.encode("utf-8")),
                    )

        # Verificar mirando todos los lugares posibles.
        verify = self._get_item(item_id)
        if self._sku_present_anywhere(verify, vid, new_sku):
            actual = self._extract_sku_from_item(verify, vid)
            print(f"  ← after PUT, actual SKU = {actual!r} ✓", flush=True)
            return result

        actual = self._extract_sku_from_item(verify, vid)
        print(f"  ← after PUT, actual SKU = {actual!r} ✗", flush=True)
        raise RuntimeError(
            f"ML aceptó el PUT (200 OK) pero NO aplicó el cambio.\n\n"
            f"SKU que intentamos setear: {new_sku!r}\n"
            f"SKU actual en ML: {actual!r}"
        )

    def _sku_present_anywhere(self, item_data: dict, variation_id, expected: str) -> bool:
        """True si `expected` aparece en cualquier campo plausible de SKU."""
        if not expected:
            return False
        if variation_id:
            try:
                vid = int(variation_id)
            except (TypeError, ValueError):
                vid = variation_id
            for v in item_data.get("variations") or []:
                if v.get("id") == vid:
                    if v.get("seller_sku") == expected:
                        return True
                    for attr in v.get("attributes") or []:
                        if (attr.get("id") or "").upper() == "SELLER_SKU" and attr.get("value_name") == expected:
                            return True
                    break
        if item_data.get("seller_custom_field") == expected:
            return True
        if item_data.get("seller_sku") == expected:
            return True
        for attr in item_data.get("attributes") or []:
            if (attr.get("id") or "").upper() == "SELLER_SKU" and attr.get("value_name") == expected:
                return True
        return False

    def _get_item(self, item_id: str) -> dict:
        url = f"https://api.mercadolibre.com/items/{item_id}"
        req = Request(
            url,
            headers={"Authorization": f"Bearer {self.auth.access_token}"},
        )
        with urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode("utf-8"))

    def _read_err_body(self, e: HTTPError) -> str:
        try:
            return e.read().decode("utf-8", errors="replace")
        except Exception:
            return ""

    def _do_put(self, url: str, body: dict):
        data = json.dumps(body).encode("utf-8")
        req = Request(
            url,
            data=data,
            method="PUT",
            headers={
                "Authorization": f"Bearer {self.auth.access_token}",
                "Content-Type": "application/json",
                "Accept": "application/json",
            },
        )
        with urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode("utf-8"))

    def _extract_sku_from_item(self, item_data: dict, variation_id) -> str:
        has_variations = bool(item_data.get("variations"))

        # 1) Variación con sus propios atributos: ese es el SKU canónico.
        if variation_id:
            try:
                vid = int(variation_id)
            except (TypeError, ValueError):
                vid = variation_id
            for v in item_data.get("variations") or []:
                if v.get("id") == vid:
                    for attr in v.get("attributes") or []:
                        if (attr.get("id") or "").upper() == "SELLER_SKU":
                            val = attr.get("value_name")
                            if val:
                                return str(val)
                    sku = v.get("seller_sku")
                    if sku:
                        return str(sku)
                    break

        # 2) Item-level. La prioridad depende de si el item tiene variaciones:
        #    - Items con variaciones: ML solo deja escribir seller_custom_field,
        #      así que ese es el campo "vivo". El attribute puede quedar viejo.
        #    - Items planos: ML usa attributes[SELLER_SKU] como canónico, y
        #      seller_custom_field puede quedar viejo si lo escribieron antes.
        if has_variations:
            sku = item_data.get("seller_custom_field")
            if sku:
                return str(sku)
            for attr in item_data.get("attributes") or []:
                if (attr.get("id") or "").upper() == "SELLER_SKU":
                    val = attr.get("value_name")
                    if val:
                        return str(val)
        else:
            for attr in item_data.get("attributes") or []:
                if (attr.get("id") or "").upper() == "SELLER_SKU":
                    val = attr.get("value_name")
                    if val:
                        return str(val)
            sku = item_data.get("seller_custom_field")
            if sku:
                return str(sku)

        sku = item_data.get("seller_sku")
        if sku:
            return str(sku)
        return ""

    def _copy_clicked_title(self):
        row = self._right_clicked_row
        if not row or self.tree.parent(row) == "":
            return
        values = self.tree.item(row, "values")
        # values = (check, time, sku, cant, producto, precio, subtotal)
        if len(values) < 5:
            return
        title = values[4]
        self._set_clipboard(title)
        self._flash_status("Título copiado ✓")

    def _open_url(self, url: str):
        """Abre una URL en el browser y trata de levantarlo al frente.

        - WSL: cmd.exe /c start (Windows ya levanta la ventana solo).
        - Linux Wayland: la política anti-focus-stealing del WM hace que
          xdg-open / Popen del binario solo "rebote" el ícono del dock.
          La forma canónica de activar una app respetando el activation
          token es `gtk-launch <desktop-id> <url>`, que pasa por GIO y
          usa el portal correctamente. Probamos varios .desktop comunes
          en orden de preferencia, con fallback a Popen del binario y por
          último a webbrowser.open."""
        import subprocess
        import shutil

        if _IS_WSL:
            try:
                subprocess.run(
                    ["cmd.exe", "/c", "start", "", url],
                    check=True,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
                return
            except (OSError, subprocess.CalledProcessError):
                pass
            webbrowser.open(url)
            return

        # Linux: gtk-launch con .desktop. Más confiable en Wayland que Popen
        # directo porque pasa por el flow de activation token de GIO.
        if shutil.which("gtk-launch"):
            for desktop_id in (
                "brave-browser",
                "com.brave.Browser",
                "google-chrome",
                "com.google.Chrome",
                "org.mozilla.firefox",
                "firefox",
            ):
                try:
                    r = subprocess.run(
                        ["gtk-launch", desktop_id, url],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        timeout=3,
                    )
                    if r.returncode == 0:
                        return
                except (OSError, subprocess.SubprocessError):
                    continue

        # Fallback 1: Popen del binario directamente.
        for browser in ("brave-browser", "google-chrome", "chromium", "firefox"):
            path = shutil.which(browser)
            if not path:
                continue
            try:
                subprocess.Popen(
                    [path, url],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    start_new_session=True,
                )
                return
            except OSError:
                continue

        # Fallback 2: webbrowser stdlib (xdg-open).
        webbrowser.open(url)

    def _set_clipboard(self, text: str):
        """Copia al clipboard. En WSL usa clip.exe (clipboard de Windows);
        en Linux nativo usa el clipboard de Tk."""
        if _IS_WSL:
            try:
                import subprocess
                subprocess.run(
                    ["clip.exe"],
                    input=text.encode("utf-16le"),
                    check=True,
                )
                return
            except (OSError, subprocess.CalledProcessError):
                pass  # fallback al clipboard de Tk
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.root.update()

    def _copy_clicked_item_id(self):
        row = self._right_clicked_row
        if not row or self.tree.parent(row) == "":
            return
        info = self.leaf_to_item.get(row) or {}
        item_id = info.get("item_id") or ""
        if not item_id:
            self._flash_status("Esta venta no tiene ID de publicación")
            return
        self._set_clipboard(item_id)
        self._flash_status("ID publicación copiado ✓")

    def _copy_clicked_sku(self):
        row = self._right_clicked_row
        if not row or self.tree.parent(row) == "":
            return
        values = self.tree.item(row, "values")
        if len(values) < 3:
            return
        sku = values[2] or ""
        if not sku:
            self._flash_status("Esta venta no tiene SKU")
            return
        self._set_clipboard(sku)
        self._flash_status("SKU copiado ✓")

    def _collect_selected(self):
        """Devuelve (ordered_days, days_dict, grand_total, count). Lee del Treeview."""
        days: dict = {}
        parent_to_day = {v: k for k, v in self.day_nodes.items()}
        grand_total = 0.0
        for parent_id in self.tree.get_children(""):
            day_key = parent_to_day.get(parent_id, "")
            for leaf_id in self.tree.get_children(parent_id):
                order_id = self.row_to_order.get(leaf_id)
                if not order_id or not local_store.is_checked(order_id):
                    continue
                values = self.tree.item(leaf_id, "values")
                _, time_str, sku, qty, title, price_str, subtotal_str = values
                info = self.leaf_to_item.get(leaf_id) or {}
                quantity = int(info.get("quantity") or qty or 1)
                unit_price = float(info.get("unit_price") or 0.0)
                line_total = float(info.get("line_total") or 0.0)
                grand_total += line_total
                days.setdefault(day_key, []).append(
                    (time_str, sku, quantity, title, unit_price, line_total)
                )

        def _key(d):
            try:
                return datetime.strptime(d, "%d/%m/%Y")
            except ValueError:
                return datetime.min

        ordered = sorted(days.keys(), key=_key, reverse=True)
        count = sum(len(days[d]) for d in ordered)
        return ordered, days, grand_total, count

    def _copy_selected_to_clipboard(self):
        if local_store.count_checked() == 0:
            self._flash_status("No hay ventas seleccionadas")
            return
        ordered, days, grand_total, total_count = self._collect_selected()
        if not days:
            self._flash_status("Las ventas seleccionadas no están cargadas")
            return

        lines = [f"*Ventas seleccionadas* ({total_count})", ""]
        for d in ordered:
            lines.append(f"📅 *{d}*")
            for time_str, sku, quantity, title, unit_price, line_total in days[d]:
                sku_part = f" [{sku}]" if sku else ""
                if quantity and quantity != 1:
                    price_part = f"{quantity} x {format_price(unit_price)} = {format_price(line_total)}"
                else:
                    price_part = format_price(unit_price)
                lines.append(f"• {time_str} — {title}{sku_part} — {price_part}")
            lines.append("")
        lines.append(f"*Total: {format_price(grand_total)}*")
        text = "\n".join(lines)

        self._set_clipboard(text)
        self._flash_status(f"Copiado al portapapeles ✓ ({total_count} ventas)")

    def export_excel(self):
        if local_store.count_checked() == 0:
            self._flash_status("No hay ventas seleccionadas")
            return
        ordered, days, grand_total, count = self._collect_selected()
        if not days:
            self._flash_status("Las ventas seleccionadas no están cargadas")
            return

        default_name = f"ventas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Exportar a Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                "Error", "Falta el módulo openpyxl. Instalalo con:\n\npip install openpyxl"
            )
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"

        thin = Side(border_style="thin", color="888888")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="DCE6F0")
        day_fill = PatternFill("solid", fgColor="F0F4F8")
        base_font = Font(size=14)
        bold = Font(bold=True, size=15)
        center = Alignment(horizontal="center", vertical="center", indent=1)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        right = Alignment(horizontal="right", vertical="center", indent=1)

        headers = ["Fecha", "Hora", "SKU", "Cant", "Producto", "Precio U.", "Subtotal", "Notas"]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = bold
            cell.alignment = center
            cell.fill = header_fill
            cell.border = border

        ncols = len(headers)
        row_idx = 2
        for d in ordered:
            for time_str, sku, quantity, title, unit_price, line_total in days[d]:
                ws.cell(row=row_idx, column=1, value=d).alignment = center
                ws.cell(row=row_idx, column=2, value=time_str).alignment = center
                ws.cell(row=row_idx, column=3, value=sku).alignment = center
                ws.cell(row=row_idx, column=4, value=quantity).alignment = center
                ws.cell(row=row_idx, column=5, value=title).alignment = left
                price_cell = ws.cell(row=row_idx, column=6, value=unit_price)
                price_cell.number_format = '"$"#,##0.00'
                price_cell.alignment = right
                subtotal_cell = ws.cell(row=row_idx, column=7, value=line_total)
                subtotal_cell.number_format = '"$"#,##0.00'
                subtotal_cell.alignment = right
                ws.cell(row=row_idx, column=8, value="")  # campo notas vacío
                for c in range(1, ncols + 1):
                    cell = ws.cell(row=row_idx, column=c)
                    cell.border = border
                    cell.font = base_font
                ws.row_dimensions[row_idx].height = 48  # más alto para escribir notas
                row_idx += 1

        # Fila de total
        ws.cell(row=row_idx, column=1, value="TOTAL").font = bold
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        merged = ws.cell(row=row_idx, column=1)
        merged.alignment = right
        total_cell = ws.cell(row=row_idx, column=7, value=grand_total)
        total_cell.font = bold
        total_cell.number_format = '"$"#,##0.00'
        total_cell.alignment = right
        ws.cell(row=row_idx, column=8, value="")
        for c in range(1, ncols + 1):
            ws.cell(row=row_idx, column=c).border = border
            ws.cell(row=row_idx, column=c).fill = day_fill

        # Anchos de columna (más amplios para fuente más grande + padding)
        widths = {1: 16, 2: 11, 3: 18, 4: 9, 5: 55, 6: 16, 7: 18, 8: 32}
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        ws.row_dimensions[1].height = 30

        # Print setup: horizontal, ajustar a 1 página de ancho
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        try:
            wb.save(path)
        except OSError as e:
            messagebox.showerror("Error", f"No se pudo guardar:\n{e}")
            return

        self._flash_status(f"Exportado: {os.path.basename(path)} ({count} ventas)")

    def _modifier_pressed(self, modifier: str):
        if modifier in self._modifiers_active:
            return  # auto-repeat de KeyPress, ignorar
        if not self._modifiers_active:
            self._status_before_hint = self.status_var.get()
        self._modifiers_active.add(modifier)
        self._refresh_modifier_hint()

    def _modifier_released(self, modifier: str):
        if modifier not in self._modifiers_active:
            return
        self._modifiers_active.discard(modifier)
        if self._modifiers_active:
            self._refresh_modifier_hint()
        else:
            if self._status_before_hint:
                self.status_var.set(self._status_before_hint)
            else:
                self._update_status()

    def _refresh_modifier_hint(self):
        mods = self._modifiers_active
        if "ctrl" in mods:
            self.status_var.set("⌨ Ctrl + click → editar precio FOB")
        elif "alt" in mods:
            self.status_var.set("⌨ Alt + click → editar multiplicador")

    def _flash_status(self, msg: str, ms: int = 2500):
        prev = self.status_var.get()
        self.status_var.set(msg)
        self.root.after(ms, lambda: self.status_var.set(prev) if self.status_var.get() == msg else None)

    def _on_click(self, event):
        # Si vienen modificadores Ctrl/Shift, los manejan los otros bindings.
        if event.state & 0x0004 or event.state & 0x0001:
            return
        # Cerrar el menú contextual si quedó abierto.
        try:
            self.context_menu.unpost()
        except tk.TclError:
            pass
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = self.tree.identify_column(event.x)
        row = self.tree.identify_row(event.y)
        if not row or col != "#1":
            return
        if self.tree.parent(row) == "":
            # Click en la columna check de un día → toggle todo el día.
            self._toggle_day(row)
        else:
            self._toggle(row)

    def refresh(self):
        if self.loading:
            return
        self.tree.delete(*self.tree.get_children())
        self.row_to_order.clear()
        self.row_base.clear()
        self.leaf_to_item.clear()
        self.day_nodes.clear()
        self.day_count.clear()
        self.day_total.clear()
        self._leaves_meta.clear()
        self._leaf_order_counter = 0
        self.offset = 0
        self._fetch_async(append=False)

    def load_more(self):
        if self.loading:
            return
        self._fetch_async(append=True)

    def _fetch_async(self, append: bool):
        self._set_loading(True)
        offset = self.offset

        def worker():
            try:
                data = fetch_orders(self.auth, offset=offset, limit=PAGE_SIZE)
            except HTTPError as e:
                self.root.after(0, self._on_error, f"HTTP {e.code}: {e.reason}")
                return
            except URLError as e:
                self.root.after(0, self._on_error, f"Error de red: {e.reason}")
                return
            except Exception as e:
                self.root.after(0, self._on_error, str(e))
                return
            self.root.after(0, self._on_data, data, append)

        threading.Thread(target=worker, daemon=True).start()

    def _on_error(self, msg: str):
        self._set_loading(False)
        messagebox.showerror("Error", msg)

    def _on_data(self, data: dict, append: bool):
        results = data.get("results", []) or []
        paging = data.get("paging", {}) or {}
        self.total = paging.get("total", len(results))

        start_idx = len(self.row_to_order)
        touched_days = set()
        for i, order in enumerate(results):
            items = order.get("order_items") or []
            first = items[0] if items else {}
            item = first.get("item") or {}
            title = item.get("title", "(sin título)")
            sku = extract_sku(first)
            unit_price = first.get("unit_price")
            try:
                quantity = int(first.get("quantity") or 1)
            except (TypeError, ValueError):
                quantity = 1
            try:
                unit_price_num = float(unit_price or 0)
            except (TypeError, ValueError):
                unit_price_num = 0.0
            line_total = unit_price_num * quantity
            price_str = format_price(unit_price)
            subtotal_str = format_price(line_total)
            dt = parse_iso(order.get("date_created", ""))
            day_key = format_day(dt)
            time_str = format_time(dt)
            order_id = str(order.get("id", ""))

            parent_id = self._get_or_create_day(day_key)
            checked = local_store.is_checked(order_id)
            mark = CHECKED if checked else UNCHECKED
            base = "odd" if (start_idx + i) % 2 else "even"
            tags_list = ["selected"] if checked else [base]
            if local_store.has_nota(order_id):
                tags_list.append("with_note")
            leaf_id = self.tree.insert(
                parent_id,
                "end",
                values=(mark, time_str, sku, quantity, title, price_str, subtotal_str),
                tags=tuple(tags_list),
            )
            self.row_to_order[leaf_id] = order_id
            self.row_base[leaf_id] = base
            # Meta para el filtro: día + texto normalizado para matchear.
            self._leaves_meta[leaf_id] = {
                "day_key": day_key,
                "row_text": _normalize(f"{sku} {title}"),
                "order": self._leaf_order_counter,
            }
            self._leaf_order_counter += 1
            # Solo guardamos lo que viene "gratis" en el listado de orders.
            # No calculamos sale_fee/taxes/shipping/neto: el neto MP se carga
            # a mano por venta (ver _open_neto_modal y local_store.neto_manual).
            payments = order.get("payments") or []
            first_payment = payments[0] if payments else {}
            try:
                total_amount = float(order.get("total_amount") or line_total)
            except (TypeError, ValueError):
                total_amount = line_total

            self.leaf_to_item[leaf_id] = {
                "item_id": item.get("id", ""),
                "variation_id": item.get("variation_id"),
                "sku": sku,
                "title": title,
                "quantity": quantity,
                "unit_price": unit_price_num,
                "line_total": line_total,
                "payment_id": first_payment.get("id"),
                "payment_method": first_payment.get("payment_method_id"),
                "total_amount": total_amount,
            }

            self.day_count[day_key] = self.day_count.get(day_key, 0) + 1
            self.day_total[day_key] = self.day_total.get(day_key, 0.0) + line_total
            touched_days.add(day_key)

        for day_key in touched_days:
            self._refresh_day_header(day_key)

        self.offset += len(results)
        self._update_header_check()
        self._set_loading(False)

        # Si hay filtro activo, aplicarlo a las filas nuevas también.
        if self._filter_active:
            self._refresh_tree_filter()

    def _get_or_create_day(self, day_key: str) -> str:
        if day_key in self.day_nodes:
            return self.day_nodes[day_key]
        parent_id = self.tree.insert(
            "",
            "end",
            text=day_key,
            values=(UNCHECKED, "", "", "", ""),
            tags=("day",),
            open=True,
        )
        self.day_nodes[day_key] = parent_id
        return parent_id

    def _refresh_day_header(self, day_key: str):
        parent_id = self.day_nodes.get(day_key)
        if not parent_id:
            return
        n = self.day_count.get(day_key, 0)
        total = self.day_total.get(day_key, 0.0)
        label = f"{day_key}  —  {n} venta{'s' if n != 1 else ''}  —  {format_price(total)}"
        self.tree.item(parent_id, text=label)
        self._update_day_check(parent_id)


def main():
    root = tk.Tk(className="ventas-ml")
    icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ventas-ml-icon.png")
    if os.path.exists(icon_path):
        icon_img = Image.open(icon_path)
        icon_sizes = []
        for size in (16, 32, 48, 64, 128, 256):
            resized = icon_img.resize((size, size), Image.LANCZOS)
            icon_sizes.append(ImageTk.PhotoImage(resized))
        root.tk.call("wm", "iconphoto", root._w, "-default", *icon_sizes)
        root._icon_refs = icon_sizes  # evitar GC
    VentasApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
